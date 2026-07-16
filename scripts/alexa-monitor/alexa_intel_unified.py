"""
alexa_intel_unified.py
统一 Alexa 问题情报：关键词 Top5 + 固定竞品 + 自有 ASIN
分四阶段构建，每阶段独立可测。
"""

import json
import math
import os
import re
import shutil
import subprocess
import sys
import time
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import sync_playwright

# ─── 配置 ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
ASINS_FILE      = SCRIPT_DIR / "asins.txt"        # 固定竞品 ASIN
OWN_ASINS_FILE  = SCRIPT_DIR / "own_asins.txt"    # 自有 ASIN
ARCHIVE_FILE    = SCRIPT_DIR / "alexa_intel_archive.json"
RAW_DIR         = SCRIPT_DIR / "raw"
REPORT_DIR      = SCRIPT_DIR / "reports"

RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d")
RUN_DATETIME  = datetime.now().strftime("%Y-%m-%d %H:%M")

SORFTIME_KEY      = "a2ngm2ruztlecldntdfnrhhretlqzz09"
SORFTIME_ENDPOINT = f"https://mcp.sorftime.com/?key={SORFTIME_KEY}"
KEYWORD_TOP_N     = 5   # 每个关键词取前 N 个自然位 ASIN

KEYWORDS = [
    {"keyword": "tripod for iphone",       "group": "ST102"},
    {"keyword": "selfie stick for iphone", "group": "ST102"},
    {"keyword": "vlogging kit for iphone", "group": "VK101"},
    {"keyword": "vlogging kit",            "group": "VK101"},
]

FEISHU_WEBHOOK = os.getenv(
    "ALEXA_FEISHU_WEBHOOK_URL",
    "https://open.feishu.cn/open-apis/bot/v2/hook/20e06d51-6ac0-4e78-8229-0dd3abd581b3",
)

# Playwright / 抓取配置
COOKIES_FILE        = SCRIPT_DIR / "amazon_cookies.json"
HEADLESS            = os.getenv("ALEXA_HEADLESS", "1").lower() not in {"0", "false", "no"}
DELAY_BETWEEN_ASINS = int(os.getenv("ALEXA_DELAY_BETWEEN_ASINS", "10"))
MAX_RETRIES         = int(os.getenv("ALEXA_MAX_RETRIES", "3"))
NAVIGATION_TIMEOUT  = int(os.getenv("ALEXA_NAVIGATION_TIMEOUT_MS", "60000"))
DOM_READY_TIMEOUT   = int(os.getenv("ALEXA_DOM_READY_TIMEOUT_MS", "15000"))
ALEXA_HINT_TIMEOUT  = int(os.getenv("ALEXA_HINT_TIMEOUT_MS", "20000"))
BLOCKED_RESOURCE_TYPES = {"image", "media", "font"}
MIN_SUCCESSFUL_ASINS   = int(os.getenv("ALEXA_MIN_SUCCESSFUL_ASINS", "15"))

# Phase 5: 本品 Alexa 问答质检
PROBE_QUESTIONS_FILE   = SCRIPT_DIR / "alexa_probe_questions.json"
QA_ENABLED             = os.getenv("ALEXA_QA_DISABLED", "0").lower() not in {"1", "true", "yes"}
QA_ANSWER_TIMEOUT_S    = int(os.getenv("ALEXA_QA_ANSWER_TIMEOUT_S", "50"))
QA_DELAY_BETWEEN_Q     = int(os.getenv("ALEXA_QA_DELAY_BETWEEN_Q", "8"))
QA_MAX_Q_PER_ASIN      = int(os.getenv("ALEXA_QA_MAX_Q_PER_ASIN", "5"))
# Alexa 面板要求高新鲜度认证（max_auth_age），静态 cookie 文件过不了，
# 必须用持久化 profile（先跑 --login-setup 人工登录一次，之后每周运行自动保鲜）
QA_PROFILE_DIR         = SCRIPT_DIR / "qa_browser_profile"

# Phase 6: 洞察合成（Python 预处理 + Claude CLI）
INSIGHT_ENABLED        = os.getenv("ALEXA_INSIGHT_DISABLED", "0").lower() not in {"1", "true", "yes"}
CLAUDE_BIN             = os.getenv("ALEXA_CLAUDE_BIN",
                                   "/Users/lihuan/.nvm/versions/node/v22.22.0/bin/claude")
LISTING_TASK_DIR       = SCRIPT_DIR.parent / "40 asin listing weekly 抓取"
LISTING_SNAPSHOT_DIR   = LISTING_TASK_DIR / "data" / "snapshots"
LISTING_DIFF_DIR       = LISTING_TASK_DIR / "data" / "raw" / "diffs"

# 赞助位哨兵（Sponsored Prompts：SP/SB campaign 自动延伸的挂件广告位，2026-03-25 GA）
OWN_BRAND              = "Vlogara"
SPONSORED_PROMPT_RE    = re.compile(r"^\s*sponsored", re.I)

# 输出路径
GITHUB_REPO       = Path("/Users/lihuan/Documents/学习提升/AI/claude/桌面 claude code/Scheduled/美日金融市场daily brief")
GITHUB_PAGES_BASE = "https://kyousan-star.github.io/-daily-brief"
GDRIVE_FOLDER_ID  = os.getenv("ALEXA_GDRIVE_FOLDER_ID", "1YCHa8JCIZjv_8FRm_QVDp2FIFUocEwtJ")
GDRIVE_TOKEN      = Path(os.getenv("ALEXA_GDRIVE_TOKEN",
                          "/Users/lihuan/.claude/scripts/gdrive_token.json")).expanduser()
OUTPUT_EXCEL = RAW_DIR   / f"alexa_intel_{RUN_TIMESTAMP}.xlsx"
OUTPUT_HTML  = REPORT_DIR / f"alexa_intel_{RUN_TIMESTAMP}.html"


# ─── Step 1: Sorftime 关键词查询 ───────────────────────────────────────────────

def _call_sorftime(name: str, arguments: dict, timeout: float = 45.0):
    payload = {
        "jsonrpc": "2.0", "id": "1",
        "method": "tools/call",
        "params": {"name": name, "arguments": arguments},
    }
    try:
        resp = requests.post(SORFTIME_ENDPOINT, json=payload, timeout=timeout)
        resp.raise_for_status()
    except Exception as e:
        print(f"    ⚠ Sorftime {name} 请求失败: {e}")
        return None

    for line in resp.text.splitlines():
        if line.startswith("data:"):
            try:
                msg = json.loads(line[5:].strip())
            except json.JSONDecodeError:
                return None
            if msg.get("result", {}).get("isError"):
                return None
            contents = msg.get("result", {}).get("content", [])
            if contents and contents[0].get("type") == "text":
                raw = contents[0]["text"]
                try:
                    return json.loads(raw)
                except json.JSONDecodeError:
                    return {"_raw": raw}
    return None


def fetch_sorftime_keyword(kw_cfg: dict) -> dict:
    """查单个关键词：搜索量/CPC + Top N 自然位 ASIN。"""
    keyword = kw_cfg["keyword"]
    print(f"  [{keyword}]")

    detail = _call_sorftime("keyword_detail",
                            {"keyword": keyword, "keywordSupportSite": "US"}) or {}
    time.sleep(1)

    page1 = _call_sorftime("keyword_search_results",
                           {"keyword": keyword, "keywordSupportSite": "US",
                            "page": 1, "positionType": 1}) or []
    time.sleep(1.5)

    def _parse_int(raw):
        try:
            return int(str(raw).replace(",", "").replace("K", "000").split(".")[0])
        except Exception:
            return 0

    def _parse_float(raw):
        try:
            return float(str(raw).replace("$", "").strip())
        except Exception:
            return 0.0

    top_n = []
    for rank_i, item in enumerate(page1[:KEYWORD_TOP_N]):
        top_n.append({
            "asin":          item.get("ASIN", ""),
            "rank":          rank_i + 1,
            "brand":         item.get("品牌", ""),
            "price":         round((item.get("价格", 0) or 0) / 100, 2),
            "monthly_sales": item.get("本产品月销量", 0) or 0,
        })

    return {
        "keyword":        keyword,
        "group":          kw_cfg["group"],
        "weekly_search":  _parse_int(detail.get("周搜索量", 0)),
        "monthly_search": _parse_int(detail.get("月搜索量", 0)),
        "cpc":            _parse_float(detail.get("推荐cpc竞价", 0)),
        "competition":    str(detail.get("搜索结果竞品数量", "")),
        "top_n":          top_n,          # list of {asin, rank, brand, price, monthly_sales}
    }


def run_sorftime_phase() -> list[dict]:
    """Phase 1：查全部关键词，返回 keyword_data 列表。"""
    print("\n═══ Phase 1: Sorftime 关键词查询 ═══════════════════════")
    results = []
    for kw_cfg in KEYWORDS:
        data = fetch_sorftime_keyword(kw_cfg)
        results.append(data)
        print(f"    周搜 {data['weekly_search']:,} | Top{KEYWORD_TOP_N}: "
              f"{[x['asin'] for x in data['top_n']]}")
    return results


# ─── Step 2: 合并 ASIN 列表，打标签 ───────────────────────────────────────────

def _load_asin_file(path: Path) -> list[str]:
    if not path.exists():
        return []
    asins = []
    seen = set()
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.split("#", 1)[0].strip()
        asin = re.sub(r"[^A-Z0-9]", "", line.upper())
        if asin and asin not in seen:
            asins.append(asin)
            seen.add(asin)
    return asins


def build_asin_roster(keyword_data: list[dict]) -> dict:
    """
    合并三个来源，去重，每个 ASIN 带完整标签。
    返回 {asin: {is_own, is_fixed, keyword_ranks: {kw: rank}}}
    优先级：own > fixed > keyword-discovered
    """
    own_asins   = set(_load_asin_file(OWN_ASINS_FILE))
    fixed_asins = set(_load_asin_file(ASINS_FILE))

    # 先把所有 ASIN 收集进来
    roster: dict[str, dict] = {}

    def _ensure(asin):
        if asin not in roster:
            roster[asin] = {"is_own": False, "is_fixed": False, "keyword_ranks": {}}

    for asin in own_asins:
        _ensure(asin)
        roster[asin]["is_own"] = True

    for asin in fixed_asins:
        _ensure(asin)
        roster[asin]["is_fixed"] = True

    for kw in keyword_data:
        for entry in kw["top_n"]:
            asin = entry["asin"]
            if not asin:
                continue
            _ensure(asin)
            roster[asin]["keyword_ranks"][kw["keyword"]] = entry["rank"]

    print(f"\n═══ Phase 2: ASIN 名单合并 ═══════════════════════════")
    total = len(roster)
    n_own     = sum(1 for v in roster.values() if v["is_own"])
    n_fixed   = sum(1 for v in roster.values() if v["is_fixed"])
    n_kw_only = sum(1 for v in roster.values()
                    if not v["is_own"] and not v["is_fixed"] and v["keyword_ranks"])
    print(f"  自有: {n_own}  固定竞品: {n_fixed}  关键词发现(新): {n_kw_only}  合计: {total}")

    return roster


# ─── Step 3: Playwright 抓取（复用 rufus_scraper 核心逻辑）──────────────────

def _normalize(text: str) -> str:
    text = re.sub(r'[\xa0​‌‍‎‏­﻿⁠]', ' ', text)
    text = unicodedata.normalize('NFC', text)
    return re.sub(r' +', ' ', text).strip()

_IGNORE_PATTERNS = [
    re.compile(r"^would you like to tell us about a lower price\?$", re.I),
    re.compile(r"^what do customers say\?$", re.I),
    re.compile(r"^what types of things can i ask\?$", re.I),
]

def _is_real_question(text: str) -> bool:
    t = _normalize(text)
    if not t.endswith('?') or len(t) < 10 or len(t) > 160:
        return False
    if any(c in t for c in '{};=><|&+*%@#^~`$\\/'):
        return False
    if re.search(r'\b(var|let|const|function|return|typeof|undefined|null\b.*null\b)\b', t):
        return False
    if not t[0].isupper() or t.count(' ') < 2:
        return False
    return True

def _filter_questions(texts: list) -> list:
    out = []
    for t in texts:
        q = _normalize(t)
        if _is_real_question(q) and not any(p.search(q) for p in _IGNORE_PATTERNS):
            out.append(q)
    return list(dict.fromkeys(out))

def _save_debug(page, asin: str, tag: str):
    try:
        debug_dir = SCRIPT_DIR / "debug"
        debug_dir.mkdir(exist_ok=True)
        stem = f"{asin}_{tag}"
        page.screenshot(path=str(debug_dir / f"{stem}.png"), full_page=False)
        (debug_dir / f"{stem}.html").write_text(page.content(), encoding="utf-8")
    except Exception:
        pass

def _extract_questions(page) -> list:
    ALEXA_SELECTORS = [
        "#dpx-nice-widget-container", "#nile-inline-btf_feature_div",
        "[data-feature-name='nile-inline']", "[data-feature-name*='alexa']",
        "[data-csa-c-slot-id*='nile']", "[data-csa-c-slot-id*='alexa']",
        "#rufus-t3", "#ask-btf_feature_div", "[data-feature-name='rufus']",
        "[id*='rufus']", "[class*='rufus']", "[id*='alexa']", "[class*='alexa']",
    ]
    for sel in ALEXA_SELECTORS:
        try:
            container = page.query_selector(sel)
            if not container:
                continue
            buttons = container.query_selector_all("button, [role='button'], span.a-size-base")
            texts = [_normalize(b.inner_text()) for b in buttons
                     if 8 < len(_normalize(b.inner_text())) < 200]
            if texts:
                print(f"    [策略1] {sel[:40]} → {len(texts)} 个")
                return texts
        except Exception:
            continue
    # 策略2：全页面以?结尾的按钮
    try:
        results = page.evaluate("""() => {
            const q = new Set();
            document.querySelectorAll('button,[role="button"],.a-button-text').forEach(el => {
                const t = (el.innerText||'').replace(/\\u00a0/g,' ').trim();
                if (t.endsWith('?') && t.length>8 && t.length<200 && !t.includes('\\n')) q.add(t);
            });
            return [...q];
        }""")
        filtered = [x for x in (results or []) if _is_real_question(x)]
        if filtered:
            print(f"    [策略2] {len(filtered)} 个")
            return filtered
    except Exception:
        pass
    return []

def _scrape_once(asin: str, context) -> dict:
    url = f"https://www.amazon.com/dp/{asin}"
    print(f"  访问: {url}")
    page = context.new_page()
    try:
        try:
            page.goto(url, timeout=NAVIGATION_TIMEOUT, wait_until="commit")
        except Exception as e:
            if "timeout" not in str(e).lower():
                raise
        try:
            page.wait_for_load_state("domcontentloaded", timeout=DOM_READY_TIMEOUT)
        except Exception:
            pass

        content_lc = page.content().lower()
        title = page.title()
        if "couldn't find that page" in content_lc or "page not found" in title.lower():
            _save_debug(page, asin, "not_found")
            return {"asin": asin, "status": "not_found", "questions": [], "error": ""}
        if "503" in title or "robot" in title.lower() or "Type the characters" in page.content()[:500]:
            _save_debug(page, asin, "blocked")
            return {"asin": asin, "status": "blocked", "questions": [], "error": title}

        # 等待+滚动触发懒加载
        try:
            page.wait_for_selector(
                "#dpx-nice-widget-container,#productTitle,#feature-bullets",
                timeout=ALEXA_HINT_TIMEOUT)
        except Exception:
            pass
        page.wait_for_timeout(2500)
        for pct in [0.15, 0.3, 0.5, 0.7, 0.9, 1.0]:
            page.evaluate(f"window.scrollTo(0, document.body.scrollHeight * {pct})")
            page.wait_for_timeout(1200)
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(2500)

        questions = _filter_questions(_extract_questions(page))
        if not questions:
            has_content = page.locator("#productTitle,#feature-bullets").count() > 0
            if not has_content:
                _save_debug(page, asin, "error")
                return {"asin": asin, "status": "timeout", "questions": [], "error": "page not loaded"}
            _save_debug(page, asin, "no_rufus")
            return {"asin": asin, "status": "no_rufus", "questions": [], "error": ""}
        print(f"    ✓ {len(questions)} 个问题")
        return {"asin": asin, "status": "ok", "questions": questions, "error": ""}
    except Exception as e:
        try:
            _save_debug(page, asin, "error")
        except Exception:
            pass
        st = "timeout" if "timeout" in str(e).lower() else "error"
        return {"asin": asin, "status": st, "questions": [], "error": str(e)}
    finally:
        try:
            page.close()
        except Exception:
            pass

def scrape_asin(asin: str, context) -> dict:
    for attempt in range(1, MAX_RETRIES + 1):
        if attempt > 1:
            print(f"    ↻ 重试 {attempt}/{MAX_RETRIES}")
            time.sleep(min(10, 2 + attempt * 2))
        result = _scrape_once(asin, context)
        if result["status"] in {"ok", "blocked", "not_found"}:
            break
    return result

def run_playwright_phase(roster: dict) -> tuple[dict, dict]:
    """
    Phase 3：抓取 roster 中所有 ASIN 的 Alexa 问题。
    返回 (asin_questions, run_statuses)
    """
    print(f"\n═══ Phase 3: Playwright 抓取（{len(roster)} 个 ASIN）══════════")
    asin_questions: dict[str, list] = {}
    run_statuses:   dict[str, dict] = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=HEADLESS,
            args=["--no-sandbox", "--disable-gpu",
                  "--disable-blink-features=AutomationControlled",
                  "--window-size=1440,900"],
        )
        context = browser.new_context(
            user_agent=("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"),
            viewport={"width": 1440, "height": 900},
            locale="en-US", timezone_id="America/New_York",
        )
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', { get: () => undefined });")
        context.route("**/*", lambda route, req: (
            route.abort() if req.resource_type in BLOCKED_RESOURCE_TYPES else route.continue_()))

        if COOKIES_FILE.exists():
            cookies = json.loads(COOKIES_FILE.read_text(encoding="utf-8"))
            valid_ss = {"Strict", "Lax", "None"}
            for c in cookies:
                if c.get("sameSite") not in valid_ss:
                    c["sameSite"] = "None"
                for k in ["hostOnly", "session", "storeId"]:
                    c.pop(k, None)
            context.add_cookies(cookies)
            print(f"✅ 已注入 {len(cookies)} 条 cookie\n")

        asins_ordered = list(roster.keys())
        for i, asin in enumerate(asins_ordered):
            meta = roster[asin]
            tag = "自有" if meta["is_own"] else ("固定竞品" if meta["is_fixed"] else "关键词新")
            print(f"[{i+1}/{len(asins_ordered)}] {asin} [{tag}]")
            result = scrape_asin(asin, context)
            run_statuses[asin]   = result
            asin_questions[asin] = result["questions"]
            if i < len(asins_ordered) - 1:
                print(f"  等待 {DELAY_BETWEEN_ASINS}s...\n")
                time.sleep(DELAY_BETWEEN_ASINS)

        try:
            browser.close()
        except Exception:
            pass

    ok_count = sum(1 for r in run_statuses.values() if r["status"] == "ok")
    print(f"\n✅ 抓取完成：{ok_count}/{len(roster)} 成功")
    return asin_questions, run_statuses


# ─── Phase 5: 本品 Alexa 问答质检 ─────────────────────────────────────────────
# 对自有 ASIN 的固定探针问题实际向 Alexa 提问，抓回答文本+建议chip，
# 分级 A明确/B含糊/C答不出/E未获取，对照 expect 极性检测 D答错，
# 并标记竞品引流chip与差评引用。2026-07-09 实测：面板 DOM 仍用 rufus-* 命名。

def load_probe_questions() -> dict:
    if not PROBE_QUESTIONS_FILE.exists():
        return {}
    try:
        raw = json.loads(PROBE_QUESTIONS_FILE.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"  ⚠️ 探针问题文件解析失败: {e}")
        return {}
    probes = {}
    for asin, items in raw.items():
        if asin.startswith("_") or not isinstance(items, list):
            continue
        cleaned = []
        for it in items[:QA_MAX_Q_PER_ASIN]:
            if isinstance(it, str):
                cleaned.append({"q": it, "expect": None})
            elif isinstance(it, dict) and it.get("q"):
                cleaned.append({"q": it["q"], "expect": it.get("expect")})
        if cleaned:
            probes[asin] = cleaned
    return probes


_HEDGE_PAT   = re.compile(r"\b(may |might |it'?s unclear|not specified|check the (product )?listing|"
                          r"contact the seller|i'?m not (sure|certain)|does not specify|unable to confirm)", re.I)
_NOINFO_PAT  = re.compile(r"(don'?t have (that |enough )?(information|details)|couldn'?t find|"
                          r"no information (available|about)|not able to find)", re.I)
_REVIEW_PAT  = re.compile(r"(customers? (report|say|mention|complain|note)|some (customers|users|buyers|reviewers)|"
                          r"reviews? (mention|indicate|suggest)|reviewers|according to (customer )?reviews)", re.I)
_NEG_PAT     = re.compile(r"^\s*no\b|\b(does not|doesn'?t|is not|isn'?t|not designed|not included|"
                          r"cannot|can'?t)\b", re.I)
_POS_PAT     = re.compile(r"^\s*yes\b|\bit (really )?does\b|\bis included\b|\bcomes with\b", re.I)


def _answer_polarity(answer: str):
    """极性只看首句（回答正文常引用五点里的 not/can't 造成全文误判）。"""
    first = re.split(r"(?<=[.!?])\s+|•", answer.strip(), maxsplit=1)[0]
    if _POS_PAT.search(first):
        return "yes"
    if _NEG_PAT.search(first):
        return "no"
    # 首句无极性再看全文，正向优先（Yes 开头被截断等情况）
    if _POS_PAT.search(answer):
        return "yes"
    if _NEG_PAT.search(answer):
        return "no"
    return None


def grade_alexa_answer(question: str, answer: str, chips: list, expect) -> dict:
    """回答分级 + 风险标记。分级看确定性，polarity 对照 expect 判答错。"""
    flags = []
    if not answer:
        return {"grade": "E", "polarity": None, "flags": ["未获取到回答"]}
    if _NOINFO_PAT.search(answer):
        grade = "C"
        polarity = None
    elif _HEDGE_PAT.search(answer):
        grade = "B"
        polarity = _answer_polarity(answer)
    else:
        grade = "A"
        polarity = _answer_polarity(answer)
    if expect and polarity and polarity != expect:
        grade = "D"
        flags.append(f"与已知事实相反（应为 {expect}）")
    if _REVIEW_PAT.search(answer):
        flags.append("⚠️ 回答引用了买家评论")
    redirect = [c for c in chips if re.match(r"^(show|find|see|browse)\b", c, re.I)]
    if redirect:
        flags.append(f"🔀 竞品引流chip: {' / '.join(redirect[:2])}")
    if grade == "C":
        flags.append("listing/属性缺该信息，Alexa 答不出")
    return {"grade": grade, "polarity": polarity, "flags": flags}


def _qa_ask_one(page, question: str) -> dict:
    """在已打开的 Alexa 面板里问一个问题，等流式回答完成后抓文本+chips。"""
    js_snapshot = """() => {
        const turns = [...document.querySelectorAll(
            '#rufus-container .rufus-papyrus-turn, #rufus-container .rufus-papyrus-active-turn')];
        const qs = [...document.querySelectorAll('#rufus-container .rufus-dialog-customer')]
            .map(q => q.textContent.replace('Customer question','').trim());
        const chips = [...document.querySelectorAll('#rufus-container button.rufus-pill')]
            .map(b => b.textContent.trim()).filter(Boolean);
        return {nQ: qs.length, nT: turns.length,
                lastTurnText: turns.length ? turns[turns.length-1].textContent : '', chips};
    }"""
    before = page.evaluate(js_snapshot)
    box = page.wait_for_selector("#rufus-text-area", timeout=10000)
    box.fill(question)
    page.keyboard.press("Enter")

    # 必须同时满足：问题气泡+1、回答轮次+1（否则会把欢迎语当回答），再等文本稳定
    deadline = time.time() + QA_ANSWER_TIMEOUT_S
    last_text, stable_since = "", None
    while time.time() < deadline:
        page.wait_for_timeout(2000)
        snap = page.evaluate(js_snapshot)
        if snap["nQ"] <= before["nQ"] or snap["nT"] <= before["nT"]:
            continue  # 回答轮次还没创建
        cur = snap["lastTurnText"]
        if cur and cur == last_text:
            if stable_since and time.time() - stable_since >= 4:
                break  # 连续两轮无变化，视为流式完成
            stable_since = stable_since or time.time()
        else:
            last_text, stable_since = cur, None
    snap = page.evaluate(js_snapshot)
    raw = snap["lastTurnText"] if snap["nT"] > before["nT"] else ""
    answer = re.sub(r"^\s*Customer question\s*", "", raw).strip()
    if answer.startswith(question):
        answer = answer[len(question):].strip()
    answer = re.sub(r"\{\}", "", answer).strip()  # React 空占位符
    return {"answer": answer, "chips": snap["chips"]}


def _qa_probe_asin(context, asin: str, probes: list) -> dict:
    """打开本品页 → 点 Ask something else 打开面板 → 逐题提问。"""
    url = f"https://www.amazon.com/dp/{asin}"
    print(f"  访问: {url}")
    page = context.new_page()
    results, status = [], "ok"
    try:
        try:
            page.goto(url, timeout=NAVIGATION_TIMEOUT, wait_until="commit")
        except Exception as e:
            if "timeout" not in str(e).lower():
                raise
        try:
            page.wait_for_load_state("domcontentloaded", timeout=DOM_READY_TIMEOUT)
        except Exception:
            pass
        title = page.title()
        if "503" in title or "robot" in title.lower():
            return {"asin": asin, "status": "blocked", "results": []}

        # 等 widget 出现并滚动触发 JS 水合（点击处理器懒加载，太早点无效）
        try:
            page.wait_for_selector("#dpx-nice-widget-container .ask-pill", timeout=ALEXA_HINT_TIMEOUT)
        except Exception:
            _save_debug(page, asin, "qa_no_widget")
            return {"asin": asin, "status": "no_widget", "results": []}
        page.evaluate("window.scrollTo(0, document.body.scrollHeight * 0.3)")
        page.wait_for_timeout(2000)
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(3000)

        # 点击开面板：处理器可能仍未挂上，最多重试 4 次
        opened = False
        for attempt in range(4):
            try:
                page.click("#dpx-nice-widget-container .ask-pill", timeout=5000)
            except Exception:
                pass
            try:
                page.wait_for_selector("#rufus-text-area", timeout=6000)
                opened = True
                break
            except Exception:
                page.wait_for_timeout(3000)
        if not opened:
            _save_debug(page, asin, "qa_panel_not_open")
            return {"asin": asin, "status": "panel_not_open", "results": []}
        page.wait_for_timeout(2000)

        # 顺路抓本品 listing 文本（title+五点+装箱清单），供洞察层做"问题×文本"交叉
        listing_text = page.evaluate("""() => {
            const t = (sel) => { const e = document.querySelector(sel); return e ? e.textContent.trim() : ''; };
            const bullets = [...document.querySelectorAll('#feature-bullets li')].map(li => li.textContent.trim());
            return {title: t('#productTitle'), bullets: bullets.slice(0, 8),
                    box: t('.postpurchase-included-components-list, #whatsInTheBoxDeck')};
        }""")

        for i, probe in enumerate(probes):
            q = probe["q"]
            print(f"    [{i+1}/{len(probes)}] 问: {q}")
            try:
                got = _qa_ask_one(page, q)
            except Exception as e:
                got = {"answer": "", "chips": []}
                print(f"      ⚠️ 提问失败: {e}")
            g = grade_alexa_answer(q, got["answer"], got["chips"], probe.get("expect"))
            print(f"      → 分级 {g['grade']}" + (f"　{'；'.join(g['flags'])}" if g["flags"] else ""))
            results.append({"question": q, "expect": probe.get("expect"),
                            "answer": got["answer"][:800], "chips": got["chips"][:6], **g})
            if i < len(probes) - 1:
                time.sleep(QA_DELAY_BETWEEN_Q)
    except Exception as e:
        status = "error"
        print(f"    ⚠️ QA 探针异常: {e}")
        try:
            _save_debug(page, asin, "qa_error")
        except Exception:
            pass
    finally:
        try:
            page.close()
        except Exception:
            pass
    out = {"asin": asin, "status": status, "results": results}
    try:
        out["listing_text"] = listing_text
    except NameError:
        out["listing_text"] = {}
    return out


def _qa_launch_context(p):
    """QA 专用持久化 profile（Alexa 面板要求高新鲜度认证，静态 cookie 过不了）。"""
    return p.chromium.launch_persistent_context(
        str(QA_PROFILE_DIR),
        headless=HEADLESS,
        args=["--no-sandbox", "--disable-gpu",
              "--disable-blink-features=AutomationControlled",
              "--window-size=1440,900"],
        viewport={"width": 1440, "height": 900},
        locale="en-US", timezone_id="America/New_York",
    )


def run_alexa_qa_phase(own_asins: set) -> dict:
    """Phase 5 入口：独立浏览器会话对自有 ASIN 跑问答质检。"""
    if not QA_ENABLED:
        print("\n↷ Phase 5 已通过 ALEXA_QA_DISABLED 跳过")
        return {}
    if not QA_PROFILE_DIR.exists():
        print("\n↷ Phase 5 跳过：QA 浏览器 profile 不存在，"
              "先运行 `python3 alexa_intel_unified.py --login-setup` 人工登录一次")
        return {}
    probes_cfg = load_probe_questions()
    targets = {a: probes_cfg[a] for a in sorted(own_asins) if a in probes_cfg}
    if not targets:
        print("\n↷ Phase 5 跳过：探针问题文件中没有匹配的自有 ASIN")
        return {}
    print(f"\n═══ Phase 5: 本品 Alexa 问答质检（{len(targets)} 个 ASIN）══════════")

    qa_results = {}
    with sync_playwright() as p:
        context = _qa_launch_context(p)
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', { get: () => undefined });")
        # 注意：QA 阶段不做资源拦截——Alexa 面板依赖动态资源，拦截可能破坏流式回答

        for i, (asin, probes) in enumerate(targets.items()):
            print(f"[{i+1}/{len(targets)}] {asin} [自有·QA]")
            qa_results[asin] = _qa_probe_asin(context, asin, probes)
            if i < len(targets) - 1:
                time.sleep(DELAY_BETWEEN_ASINS)
        try:
            context.close()
        except Exception:
            pass

    n_ans = sum(1 for r in qa_results.values() for x in r["results"] if x["grade"] != "E")
    n_all = sum(len(r["results"]) for r in qa_results.values())
    print(f"\n✅ 问答质检完成：{n_ans}/{n_all} 个问题拿到回答")
    return qa_results


# ─── Phase 6: 洞察合成 ────────────────────────────────────────────────────────
# 目标：①如何用 Alexa 优化本品 listing 让 AI 更愿意推荐 ②竞手动态/漏洞/可借鉴。
# Python 侧做事实预处理（问题×listing文本交叉、竞手listing变更、widget问题增减），
# Claude CLI 按固化的分析框架合成行动建议。信号不足时输出空，不编造。

# 问题→答案关键词词典（判断 listing 文本是否包含某高频问题的答案素材）
_COVERAGE_TERMS = [
    (re.compile(r"carry|case|bag", re.I),          ["carry", "case", " bag", "pouch", "storage"]),
    (re.compile(r"remote.*(recharge|batter)|(recharge|batter).*remote", re.I),
                                                    ["rechargeable remote", "remote control works", "cr2032", "usb"]),
    (re.compile(r"\bandroid\b", re.I),              ["android"]),
    (re.compile(r"tablet|ipad", re.I),              ["tablet", "ipad"]),
    (re.compile(r"tall|height|extend", re.I),       ["inch", "extends", " tall", "height"]),
    (re.compile(r"stable|wobble|sturdy", re.I),     ["stab", "sturdy", "anti-slip", "wobble", "secure"]),
    (re.compile(r"set ?up|install|assembl", re.I),  ["setup", "set up", "no tool", "plug", "easy to"]),
    (re.compile(r"noise|microphone|mic\b", re.I),   ["noise", "mic", "wind"]),
    (re.compile(r"light|brightness|led", re.I),     ["light", "brightness", "led", "dimmable"]),
    (re.compile(r"video call|webcam|zoom", re.I),   ["video call", "webcam", "zoom", "facetime"]),
    (re.compile(r"vertical|horizontal|portrait|landscape|rotat", re.I),
                                                    ["portrait", "landscape", "vertical", "horizontal", "360"]),
    (re.compile(r"weight|capacity|hold|heavy", re.I), [" lb", "weight", "capacity", "load", "oz"]),
    (re.compile(r"water|weather", re.I),            ["waterproof", "weather", "splash"]),
    (re.compile(r"iphone|phone.*compatib|compatib.*phone", re.I), ["iphone", "compatib", "fits phones"]),
]


def _question_coverage(question: str, listing_text: str):
    """返回 True(文本含答案素材)/False(不含)/None(问题类别不认识，不判)。"""
    txt = (listing_text or "").lower()
    for q_pat, terms in _COVERAGE_TERMS:
        if q_pat.search(question):
            return any(t.lower() in txt for t in terms)
    return None


def _latest_file(d: Path, pattern: str):
    files = sorted(d.glob(pattern)) if d.exists() else []
    return files[-1] if files else None


def _load_listing_snapshot() -> dict:
    """任务#10 当天 02:00 的快照：asin → {brand,title,bullets,...}（02:35 跑正好接上）"""
    f = _latest_file(LISTING_SNAPSHOT_DIR, "snapshot_*.json")
    if not f:
        return {}
    try:
        items = json.loads(f.read_text(encoding="utf-8")).get("items", [])
    except Exception:
        return {}
    out = {}
    for it in items:
        bullets = it.get("bullet_points") or ""
        if isinstance(bullets, list):
            bullets = " ".join(str(b) for b in bullets)
        out[it.get("asin")] = {
            "brand": it.get("brand") or "",
            "title": (it.get("title") or "")[:220],
            "text": ((it.get("title") or "") + " " + str(bullets))[:2200],
            "review_count": it.get("review_count"),
            "bsr_rank": it.get("bsr_rank"),
        }
    return out


def _load_listing_diff() -> list:
    """任务#10 的周 diff，过滤出 listing 实质变更（title/五点/价格/主图/A+/变体）。"""
    f = _latest_file(LISTING_DIFF_DIR, "weekly_diff_*.json")
    if not f:
        return []
    try:
        diff = json.loads(f.read_text(encoding="utf-8"))
    except Exception:
        return []
    MEANINGFUL = {"title", "bullet_points", "price", "main_image", "aplus_image_count",
                  "aplus_main_image_count", "variation_asins", "brand"}
    moves = []
    for ch in diff.get("changes", []):
        fields = [c for c in ch.get("changes", []) if c.get("field") in MEANINGFUL]
        if fields:
            moves.append({"asin": ch["asin"],
                          "fields": [{k: v for k, v in c.items()
                                      if k in {"field", "old", "new", "old_count", "new_count", "added", "removed"}}
                                     for c in fields]})
    return moves


def build_insight_facts(current: dict, wow: dict, prev: dict | None) -> dict:
    """组装给 Claude 的紧凑事实包（控制在 ~12KB 内）。"""
    snapshot = _load_listing_snapshot()
    roster = current["roster"]

    # 1) 本品 QA + 本品 listing 覆盖
    own_qa_facts = {}
    for asin, item in current.get("own_qa", {}).items():
        lt = item.get("listing_text", {}) or {}
        own_text = " ".join([lt.get("title", ""), " ".join(lt.get("bullets", [])), lt.get("box", "")])
        own_qa_facts[asin] = {
            "listing_title": lt.get("title", "")[:200],
            "qa": [{"q": r["question"], "grade": r["grade"], "polarity": r["polarity"],
                    "expect": r.get("expect"), "flags": r["flags"],
                    "answer": r["answer"][:350], "chips": r["chips"][:4]}
                   for r in item["results"]],
        }
        # 本品对竞品高频问题的文本覆盖
        cov = {}
        for q, asins in current["comp_analysis"].get("high_freq", []):
            c = _question_coverage(q, own_text)
            if c is not None:
                cov[q] = c
        own_qa_facts[asin]["own_text_covers_hot_questions"] = cov

    # 2) 竞手侧：重点竞品（进关键词Top5自然位的 + 评论数最高的固定竞品）
    key_asins = [a for a, m in roster.items() if m.get("keyword_ranks") and not m.get("is_own")]
    by_reviews = sorted((a for a, m in roster.items() if m.get("is_fixed") and a in snapshot),
                        key=lambda a: -(int(snapshot[a].get("review_count") or 0)))
    focus = list(dict.fromkeys(key_asins + by_reviews[:6]))[:12]

    hot_qs = [q for q, _ in current["comp_analysis"].get("high_freq", [])]
    for kw in current["keyword_analysis"]:
        for q, _ in kw["analysis"].get("high_freq", []):
            if q not in hot_qs:
                hot_qs.append(q)
    hot_qs = hot_qs[:10]

    comp_facts = []
    for a in focus:
        s = snapshot.get(a, {})
        cov = {}
        for q in hot_qs:
            c = _question_coverage(q, s.get("text", ""))
            if c is not None:
                cov[q] = c
        comp_facts.append({
            "asin": a, "brand": s.get("brand", "?"), "title": s.get("title", "")[:120],
            "keyword_ranks": roster.get(a, {}).get("keyword_ranks", {}),
            "widget_questions": current["all_questions"].get(a, [])[:6],
            "text_covers_hot_questions": cov,
        })

    # 3) 竞手动态：listing 变更 + widget 问题增减；本品 listing 变更单独拆出，不与竞品动态混在一起
    all_listing_moves = [mv for mv in _load_listing_diff() if mv["asin"] in roster]
    own_listing_moves = [mv for mv in all_listing_moves if roster.get(mv["asin"], {}).get("is_own")]
    listing_moves = [mv for mv in all_listing_moves if not roster.get(mv["asin"], {}).get("is_own")][:12]
    q_moves = []
    if prev:
        prev_q = prev.get("competitor_questions", {})
        for a, qs in current["all_questions"].items():
            if not roster.get(a, {}).get("is_fixed"):
                continue
            pq = set(prev_q.get(a, []))
            if not pq:
                continue
            added, removed = sorted(set(qs) - pq), sorted(pq - set(qs))
            if added or removed:
                q_moves.append({"asin": a, "brand": snapshot.get(a, {}).get("brand", "?"),
                                "questions_added": added[:4], "questions_removed": removed[:4]})

    return {
        "run_date": current["run_date"],
        "own_products": own_qa_facts,
        "own_gaps": {kw["keyword"]: {oa: g[:5] for oa, g in kw["own_gap"].items()}
                     for kw in current["keyword_analysis"] if kw["own_gap"]},
        "competitor_high_freq": [(q, len(asins)) for q, asins in
                                 current["comp_analysis"].get("high_freq", [])],
        "competitors": comp_facts,
        "own_listing_changes_this_week": own_listing_moves,
        "competitor_listing_changes_this_week": listing_moves,
        "competitor_widget_question_changes": q_moves[:10],
        "qa_grade_changes_vs_last_week": wow.get("qa_wow", []) if wow else [],
    }


_INSIGHT_PROMPT = """你是亚马逊 Alexa（站内AI购物助手，原Rufus）优化顾问。基于下方 JSON 事实包，为 Vlogara 品牌（自有 ASIN 见 own_products）产出本周洞察。服务两个目标：
①让 Alexa 更愿意推荐本品：Alexa 回答质量由 listing 文本+后台属性字段决定，回答得越明确越有利转化与推荐；
②竞手情报：竞手 listing 动态解读、竞手漏洞（=我的攻击点）、竞手做得好的（=可借鉴）。

分析原则（必须遵守）：
- widget 问题是 Alexa 引导买家问的"考试大纲"；本品 QA 分级 D=Alexa 把事实答错（优先查后台属性字段），C=listing 缺信息（补文案），B=表述含糊（改写更明确），A=良好；
- 区分「真实功能缺失」和「listing 表述缺失」：功能缺失改文案没用，给话术对冲或产品迭代建议；表述缺失给具体落点（标题/五点/A+/属性字段/QA）；
- 本品被 Alexa 否定回答且出现 Show xxx 引流 chip = 买家被一键带去竞品，量化损失点；
- 竞手高频被问但其 listing 文本未覆盖（text_covers_hot_questions 为 false）= 竞手漏洞，我方若覆盖了就是广告语/对比图/A+ 的攻击素材；竞手覆盖好的写法 = 可借鉴；
- 竞手 listing 变更（标题/五点/主图/价格/A+/变体）+ widget 问题增减 = 竞手运营动作，解读意图；
- own_listing_changes_this_week = 本品自己这周被谁改动了（标题/五点/主图/价格/A+/变体）；这是我方动作不是竞品动作，重点看它是否解决/引入了 QA 分级或 Gap 里的问题（对照 qa_grade_changes_vs_last_week、own_gaps），而不是当竞品情报处理；
- 事实包里没有的信息不要编；信号不足就少写或写空数组，宁缺毋滥。

只输出 JSON（不要 markdown 代码块），结构：
{"summary_line": "一句话本周结论(≤40字)",
 "actions": [{"priority": "P0|P1|P2", "signal": "触发信号(引用具体数据)", "insight": "为什么重要", "action": "具体动作", "target": "标题|五点|A+|后台属性|QA|广告|产品决策"}],
 "own_listing_moves": [{"asin": "", "move": "变了什么", "read": "是否解决/引入了QA或Gap问题，对Alexa回答质量的影响"}],
 "competitor_moves": [{"brand": "", "asin": "", "move": "变了什么", "read": "意图解读+对我影响"}],
 "attack_points": [{"question": "", "weak_competitors": ["brand"], "my_coverage": true, "how_to_use": "怎么打"}],
 "learn_from": [{"brand": "", "what": "做得好的点", "apply": "怎么借鉴"}]}
actions 最多5条按优先级排序；own_listing_moves 事实包为空数组时也输出空数组，不要编；competitor_moves/attack_points/learn_from 各最多4条。competitor_moves 的 brand 字段不能为空——事实包里查不到品牌名就填 ASIN。所有文本字段写完整句子，不要写半句。

事实包：
"""


def run_insight_phase(facts: dict) -> dict:
    """调 Claude CLI 合成洞察。失败/超时返回空 dict，不阻塞报告生成。"""
    if not INSIGHT_ENABLED:
        print("\n↷ Phase 6 已通过 ALEXA_INSIGHT_DISABLED 跳过")
        return {}
    if not Path(CLAUDE_BIN).exists():
        print(f"\n↷ Phase 6 跳过：找不到 claude CLI（{CLAUDE_BIN}）")
        return {}
    print("\n═══ Phase 6: 洞察合成（Claude CLI）══════════════════════")
    prompt = _INSIGHT_PROMPT + json.dumps(facts, ensure_ascii=False)
    try:
        result = subprocess.run([CLAUDE_BIN, "-p", prompt],
                                capture_output=True, text=True, timeout=420)
        raw = result.stdout.strip()
        if result.returncode != 0:
            print(f"  ⚠️ claude 退出码 {result.returncode}: {result.stderr[:200]}")
            return {}
        # 容错：剥掉可能的 ```json 围栏，取第一个 { 到最后一个 }
        m = re.search(r"\{.*\}", raw, re.S)
        insights = json.loads(m.group(0)) if m else {}
        n_act = len(insights.get("actions", []))
        print(f"✅ 洞察合成完成：{n_act} 条行动建议，"
              f"{len(insights.get('competitor_moves', []))} 条竞手动态，"
              f"{len(insights.get('attack_points', []))} 个攻击点")
        return insights
    except subprocess.TimeoutExpired:
        print("  ⚠️ claude 调用超时（420s），跳过洞察")
        return {}
    except Exception as e:
        print(f"  ⚠️ 洞察合成失败: {e}")
        return {}


def main_login_setup():
    """打开 QA 专用浏览器窗口让用户人工登录 Amazon（脚本不接触任何密码）。
    登录状态存进持久化 profile，之后 Phase 5 每周运行自动保鲜。"""
    print("═══ QA 浏览器 profile 登录设置 ══════════════════════════")
    print("即将打开浏览器窗口，请在窗口里手动登录 Amazon（建议用 QA 专用买家账号）。")
    print("登录完成后脚本会自动检测并验证 Alexa 面板是否可用。\n")
    QA_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            str(QA_PROFILE_DIR), headless=False,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled",
                  "--window-size=1440,900"],
            viewport={"width": 1440, "height": 900},
            locale="en-US", timezone_id="America/New_York",
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.goto("https://www.amazon.com/", timeout=60000, wait_until="domcontentloaded")
        print("⏳ 等待登录（最长 5 分钟）……请在弹出的窗口操作")
        signed_in = False
        for _ in range(100):  # 100 × 3s = 5min
            try:
                txt = page.evaluate(
                    "() => (document.querySelector('#nav-link-accountList-nav-line-1')||{textContent:''}).textContent")
                if txt and "sign in" not in txt.lower():
                    signed_in = True
                    break
            except Exception:
                pass
            time.sleep(3)
        if not signed_in:
            print("✗ 5 分钟内未检测到登录，退出。可重新运行 --login-setup")
            context.close()
            return
        print(f"✅ 已登录（{txt.strip()}）。验证 Alexa 面板……")
        page.goto("https://www.amazon.com/dp/B0GY7Y6C63", timeout=60000, wait_until="domcontentloaded")
        page.wait_for_timeout(5000)
        try:
            page.click("#nav-rufus-disco", timeout=8000)
            page.wait_for_selector("#rufus-text-area", timeout=15000)
            print("✅ Alexa 面板可用，Phase 5 就绪。窗口即将关闭。")
        except Exception:
            print("⚠️ Alexa 面板未能打开——登录可能成功但面板仍受限，跑一次 --qa-only 再确认")
        page.wait_for_timeout(2000)
        context.close()


# ─── Step 4: 三层分析 ─────────────────────────────────────────────────────────

def _analyze_question_set(asin_questions: dict) -> dict:
    """通用问题频率分析（高频/中频/独有）。"""
    total = len(asin_questions)
    q2a: dict[str, list] = defaultdict(list)
    for asin, qs in asin_questions.items():
        for q in qs:
            if asin not in q2a[q]:
                q2a[q].append(asin)
    threshold = max(2, math.ceil(total * 0.35))
    high_freq = sorted([(q, a) for q, a in q2a.items() if len(a) >= threshold],
                       key=lambda x: len(x[1]), reverse=True)
    mid_freq  = sorted([(q, a) for q, a in q2a.items() if 2 <= len(a) < threshold],
                       key=lambda x: len(x[1]), reverse=True)
    unique_qs = {asin: [q for q in qs if len(q2a[q]) == 1]
                 for asin, qs in asin_questions.items()}
    return {"total": total, "threshold": threshold, "q2a": dict(q2a),
            "high_freq": high_freq, "mid_freq": mid_freq, "unique": unique_qs}


def run_analysis_phase(keyword_data: list[dict], roster: dict,
                       asin_questions: dict, run_statuses: dict) -> dict:
    """
    Phase 4：三层交叉分析。
    返回 analysis_result dict，供 HTML/Feishu/Excel 使用。
    """
    print("\n═══ Phase 4: 三层分析 ══════════════════════════════════")

    own_asins    = {a for a, m in roster.items() if m["is_own"]}
    fixed_asins  = {a for a, m in roster.items() if m["is_fixed"]}

    # ── 只统计成功抓取的 ASIN ──
    ok_questions = {a: qs for a, qs in asin_questions.items()
                    if run_statuses.get(a, {}).get("status") == "ok" and qs}

    # ── 赞助位哨兵：把 Sponsored Prompts 广告位从有机问题中剥离 ──
    # （广告位混进高频/gap 分析会污染口径；剥离后单独分类判定）
    sponsored_hits = []
    for a in list(ok_questions):
        organic = []
        for q in ok_questions[a]:
            if SPONSORED_PROMPT_RE.match(q):
                sponsored_hits.append({"asin": a, "prompt": q.strip()})
            else:
                organic.append(q)
        ok_questions[a] = organic

    # 宿主品牌查找：任务#10快照 + 关键词 Top5 明细
    snapshot = _load_listing_snapshot()
    brand_lookup = {a: v["brand"] for a, v in snapshot.items() if v.get("brand")}
    for kw in keyword_data:
        for e in kw["top_n"]:
            if e.get("asin") and e.get("brand"):
                brand_lookup.setdefault(e["asin"], e["brand"])

    known_brands = sorted({b for b in brand_lookup.values() if b}, key=len, reverse=True)
    for h in sponsored_hits:
        host = brand_lookup.get(h["asin"], "")
        p = h["prompt"].lower()
        h["host_brand"] = host
        h["brand"] = next((b for b in known_brands if b.lower() in p), "")
        if not h["brand"]:
            mm = re.search(r"why choose\s+([\w&'+.-]+)", h["prompt"], re.I)
            h["brand"] = mm.group(1) if mm else ""
        if a := h["asin"]:
            if a in own_asins:
                h["cls"] = "own_enrolled" if OWN_BRAND.lower() in p else "invaded"
            elif host and host.lower() in p:
                h["cls"] = "self"
            elif host:
                h["cls"] = "cross"
            else:
                h["cls"] = "unknown"
    n_abnormal = sum(1 for h in sponsored_hits if h["cls"] in ("invaded", "cross"))
    print(f"  赞助位哨兵：{len(sponsored_hits)} 条，异常 {n_abnormal} 条")
    sponsored_watch = {"hits": sponsored_hits}

    # ── 竞品层（固定40） ──
    competitor_q = {a: ok_questions[a] for a in fixed_asins if a in ok_questions}
    comp_analysis = _analyze_question_set(competitor_q)
    print(f"  竞品层：{len(competitor_q)}/{len(fixed_asins)} 个成功，"
          f"高频 {len(comp_analysis['high_freq'])} 条")

    # ── 自有 ASIN 层 ──
    own_q = {a: ok_questions.get(a, []) for a in own_asins}
    own_status = {a: run_statuses.get(a, {}).get("status", "unknown") for a in own_asins}

    # ── 关键词层：每个关键词 Top5 的问题频率 ──
    keyword_analysis = []
    for kw in keyword_data:
        top5_asins = [e["asin"] for e in kw["top_n"] if e["asin"]]
        top5_ok_q  = {a: ok_questions[a] for a in top5_asins if a in ok_questions}
        kw_analysis = _analyze_question_set(top5_ok_q) if top5_ok_q else {}

        # 自有 ASIN gap：Top5 高频问题中，自有 ASIN 没有的
        top5_hf_questions = {q for q, _ in kw_analysis.get("high_freq", [])}
        own_gap: dict[str, list] = {}
        for own_asin in own_asins:
            own_qs_set = set(ok_questions.get(own_asin, []))
            gap = sorted(top5_hf_questions - own_qs_set)
            if gap:
                own_gap[own_asin] = gap

        # 自有 ASIN 在本关键词的自然位
        own_rank_here = {a: roster[a]["keyword_ranks"].get(kw["keyword"])
                         for a in own_asins if kw["keyword"] in roster[a]["keyword_ranks"]}

        keyword_analysis.append({
            "keyword":        kw["keyword"],
            "group":          kw["group"],
            "weekly_search":  kw["weekly_search"],
            "monthly_search": kw["monthly_search"],
            "cpc":            kw["cpc"],
            "competition":    kw["competition"],
            "top5":           kw["top_n"],
            "top5_ok_count":  len(top5_ok_q),
            "analysis":       kw_analysis,
            "own_gap":        own_gap,
            "own_rank":       own_rank_here,
        })
        hf_count = len(kw_analysis.get("high_freq", []))
        print(f"  [{kw['keyword']}] Top5成功={len(top5_ok_q)}, 高频={hf_count}")

    return {
        "run_date":         RUN_DATETIME,
        "keyword_analysis": keyword_analysis,
        "comp_analysis":    comp_analysis,
        "own_q":            own_q,
        "own_status":       own_status,
        "own_asins":        sorted(own_asins),
        "all_questions":    ok_questions,
        "run_statuses":     run_statuses,
        "roster":           roster,
        "sponsored_watch":  sponsored_watch,
    }


# ─── Step 4b: WoW 对比 ────────────────────────────────────────────────────────

def load_prev_archive() -> dict | None:
    if not ARCHIVE_FILE.exists():
        return None
    try:
        return json.loads(ARCHIVE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return None


def compute_wow(current: dict, prev: dict) -> dict:
    """对比竞品层和关键词层的 WoW 变化。"""
    if not prev:
        return {}

    # 竞品层 WoW
    prev_comp_q   = prev.get("competitor_questions", {})
    curr_comp_q   = current["all_questions"]
    prev_all      = set(q for qs in prev_comp_q.values() for q in qs)
    curr_all      = set(q for qs in curr_comp_q.values() for q in qs)

    prev_comp_ana = _analyze_question_set(prev_comp_q) if prev_comp_q else {}
    curr_comp_ana = current["comp_analysis"]
    prev_hf = {q for q, _ in prev_comp_ana.get("high_freq", [])}
    curr_hf = {q for q, _ in curr_comp_ana.get("high_freq", [])}

    # 关键词层 WoW（搜索量+关键词高频问题变化）
    prev_kw_map = {kw["keyword"]: kw for kw in prev.get("keyword_analysis", [])}
    kw_wow = []
    for kw in current["keyword_analysis"]:
        prev_kw = prev_kw_map.get(kw["keyword"])
        if not prev_kw:
            kw_wow.append({"keyword": kw["keyword"], "has_prev": False})
            continue
        weekly_delta  = kw["weekly_search"]  - prev_kw.get("weekly_search", 0)
        monthly_delta = kw["monthly_search"] - prev_kw.get("monthly_search", 0)
        cpc_delta     = round(kw["cpc"] - prev_kw.get("cpc", 0), 2)
        prev_kw_hf    = {q for q, _ in prev_kw.get("top5_high_freq", [])}
        curr_kw_hf    = {q for q, _ in kw["analysis"].get("high_freq", [])}
        kw_wow.append({
            "keyword":        kw["keyword"],
            "has_prev":       True,
            "weekly_delta":   weekly_delta,
            "monthly_delta":  monthly_delta,
            "cpc_delta":      cpc_delta,
            "kw_hf_new":      sorted(curr_kw_hf - prev_kw_hf),
            "kw_hf_dropped":  sorted(prev_kw_hf - curr_kw_hf),
        })

    # 问答质检 WoW：分级变化即报警信号
    qa_wow = []
    prev_qa = prev.get("own_qa", {})
    for asin, item in current.get("own_qa", {}).items():
        prev_map = {r["question"]: r for r in prev_qa.get(asin, [])}
        for r in item["results"]:
            pr = prev_map.get(r["question"])
            if pr and pr.get("grade") != r["grade"]:
                qa_wow.append({"asin": asin, "question": r["question"],
                               "prev_grade": pr.get("grade"), "curr_grade": r["grade"]})

    # 赞助位：品牌首次进场（对照历史累计名单，抗轮播抽样波动）
    curr_sp_brands = {h["brand"] for h in current.get("sponsored_watch", {}).get("hits", [])
                      if h.get("brand") and h["brand"].lower() != OWN_BRAND.lower()}
    prev_seen = set(prev.get("sponsored_brands_seen", []))
    sponsored_new_brands = sorted(curr_sp_brands - prev_seen) if prev_seen else []
    current["sponsored_brands_seen"] = sorted(prev_seen | curr_sp_brands)

    return {
        "prev_date":      prev.get("run_date", "上周"),
        "sponsored_new_brands": sponsored_new_brands,
        "comp_hf_new":    sorted(curr_hf - prev_hf),
        "comp_hf_drop":   sorted(prev_hf - curr_hf),
        "brand_new_q":    sorted(curr_all - prev_all),
        "disappeared_q":  sorted(prev_all - curr_all),
        "kw_wow":         kw_wow,
        "qa_wow":         qa_wow,
    }


def save_archive(current: dict):
    """存档本次结果，供下周 WoW 使用。"""
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    data = {
        "run_date":             current["run_date"],
        "keyword_analysis":     [
            {
                "keyword":          kw["keyword"],
                "group":            kw["group"],
                "weekly_search":    kw["weekly_search"],
                "monthly_search":   kw["monthly_search"],
                "cpc":              kw["cpc"],
                "top5_high_freq":   [(q, len(a)) for q, a in kw["analysis"].get("high_freq", [])],
            }
            for kw in current["keyword_analysis"]
        ],
        "competitor_questions": {a: qs for a, qs in current["all_questions"].items()
                                 if current["roster"].get(a, {}).get("is_fixed")},
        "own_questions":        current["own_q"],
        "own_qa": {
            asin: [{"question": r["question"], "grade": r["grade"],
                    "polarity": r["polarity"], "flags": r["flags"]}
                   for r in item["results"]]
            for asin, item in current.get("own_qa", {}).items()
        },
        "sponsored_prompts": current.get("sponsored_watch", {}).get("hits", []),
        "sponsored_brands_seen": current.get("sponsored_brands_seen")
            or sorted({h["brand"] for h in current.get("sponsored_watch", {}).get("hits", [])
                       if h.get("brand")}),
    }
    ARCHIVE_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✅ 存档已保存: {ARCHIVE_FILE}")


# ─── Step 5: 输出 ─────────────────────────────────────────────────────────────

def _esc(s) -> str:
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


# ── Gap × 质检交叉：gap 问题若质检已 A 级明确作答，说明仅挂件未展示，不是内容缺失 ──

def _norm_q(s: str) -> str:
    return "".join(c for c in s.lower() if c.isalnum() or c == " ").strip()

def _build_qa_lookup(current: dict) -> dict:
    return {asin: {_norm_q(r["question"]): r["grade"] for r in item["results"]}
            for asin, item in current.get("own_qa", {}).items()}

def _gap_grade(qa_lookup: dict, asin: str, q: str):
    nq = _norm_q(q)
    for pq, g in qa_lookup.get(asin, {}).items():
        if nq == pq or nq in pq or pq in nq:
            return g
    return None

def generate_html(current: dict, wow: dict) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    ka   = current["keyword_analysis"]
    ca   = current["comp_analysis"]
    own  = current["own_q"]
    own_statuses = current["own_status"]
    roster = current["roster"]

    CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif; background:#f5f7fa; color:#2c3e50; }
.hdr { background: linear-gradient(135deg,#1a237e,#283593); color:white; padding:20px 32px; }
.hdr h1 { font-size:18px; font-weight:600; margin-bottom:4px; }
.hdr .sub { font-size:12px; opacity:.85; }
.wrap { max-width:1100px; margin:0 auto; padding:20px 16px; }
.sec { background:white; border-radius:10px; padding:18px 22px; margin-bottom:18px;
       box-shadow:0 1px 4px rgba(0,0,0,.08); }
.sec h2 { font-size:14px; font-weight:600; color:#1a237e; border-bottom:1px solid #e8ecf0;
          padding-bottom:8px; margin-bottom:14px; }
.sec h3 { font-size:13px; font-weight:600; color:#283593; margin:12px 0 6px; }
table { width:100%; border-collapse:collapse; font-size:12px; }
th { background:#f0f2f8; text-align:left; padding:7px 10px; color:#333; font-weight:600; }
td { padding:6px 10px; border-bottom:1px solid #f0f2f0; vertical-align:top; }
tr:last-child td { border-bottom:none; }
.bar-wrap { background:#eef0f8; border-radius:3px; height:5px; width:80px;
            display:inline-block; vertical-align:middle; }
.bar { background:#3949ab; height:5px; border-radius:3px; }
.tag-h { background:#e8eaf6; color:#3949ab; padding:1px 6px; border-radius:3px; font-size:11px; font-weight:600; }
.tag-m { background:#e8f5e9; color:#2e7d32; padding:1px 6px; border-radius:3px; font-size:11px; }
.tag-u { background:#fce4ec; color:#c62828; padding:1px 6px; border-radius:3px; font-size:11px; }
.wn { color:#2e7d32; font-weight:600; }
.wd { color:#c62828; }
.wow-box { background:#fffde7; border-left:4px solid #f9a825; padding:12px 16px; border-radius:5px; margin-bottom:12px; }
.kw-card { border:1px solid #e0e4f0; border-radius:8px; padding:14px 16px; margin-bottom:12px; }
.kw-title { font-size:13px; font-weight:700; color:#1a237e; }
.kw-meta { font-size:12px; color:#666; margin:4px 0 10px; }
.asin-hd { background:#f0f2f8; padding:3px 8px; border-radius:4px; font-size:12px; font-weight:600; display:inline-block; margin:6px 0 4px; }
.qi { font-size:12px; padding:3px 0 3px 10px; border-left:2px solid #e0e0e0; margin:2px 0; }
.gap-box { background:#fff3e0; border-left:3px solid #ff9800; padding:8px 12px; border-radius:4px; margin-top:8px; font-size:12px; }
.gap-ok { background:#f1f8e9; border-left:3px solid #8bc34a; padding:8px 12px; border-radius:4px; margin-top:8px; font-size:12px; }
.legend { font-size:11px; color:#999; margin:-8px 0 12px; }
details summary { cursor:pointer; font-size:12px; font-weight:600; color:#3949ab; padding:6px 0; }
"""
    parts = [f"""<!DOCTYPE html><html lang="zh"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Alexa 综合情报 {_esc(RUN_TIMESTAMP)}</title><style>{CSS}</style></head>
<body>
<div class="hdr">
  <h1>📊 Alexa 问题综合情报 — 手机三脚架/Vlogging Kit</h1>
  <div class="sub">生成时间：{_esc(RUN_DATETIME)} &nbsp;|&nbsp;
  共抓取 {len(current['all_questions'])} 个ASIN
  {"&nbsp;|&nbsp; 对比基准：" + _esc(wow.get("prev_date","")) if wow else ""}</div>
</div>
<div class="wrap">"""]

    # ── 洞察层（Phase 6）：行动建议 + 竞手动态置顶，决策先行 ──
    ins = current.get("insights", {}) or {}
    if ins.get("summary_line"):
        parts.append(f'<div class="sec" style="border-left:4px solid #1a237e">'
                     f'<div style="font-size:14px;font-weight:700;color:#1a237e">📌 {_esc(ins["summary_line"])}</div></div>')
    if ins.get("actions"):
        PRI_CLS = {"P0": "tag-u", "P1": "tag-h", "P2": "tag-m"}
        parts.append('<div class="sec"><h2>🎯 本周行动建议</h2>')
        parts.append('<table><tr><th style="width:6%">优先级</th><th style="width:22%">触发信号</th>'
                     '<th style="width:26%">解读</th><th>建议动作</th><th style="width:10%">落点</th></tr>')
        for a in ins["actions"]:
            cls = PRI_CLS.get(a.get("priority", "P2"), "tag-m")
            parts.append(f'<tr><td><span class="{cls}">{_esc(a.get("priority",""))}</span></td>'
                         f'<td>{_esc(a.get("signal",""))}</td><td>{_esc(a.get("insight",""))}</td>'
                         f'<td>{_esc(a.get("action",""))}</td><td>{_esc(a.get("target",""))}</td></tr>')
        parts.append('</table></div>')
    if ins.get("own_listing_moves"):
        parts.append('<div class="sec" style="border-left:4px solid #2e7d32"><h2>🏠 本品 Listing 本周变化</h2>')
        parts.append('<table><tr><th style="width:12%">ASIN</th><th style="width:34%">变更</th><th>解读（是否解决/引入 QA·Gap 问题）</th></tr>')
        for mv in ins["own_listing_moves"]:
            parts.append(f'<tr><td>{_esc(mv.get("asin",""))}</td>'
                         f'<td>{_esc(mv.get("move",""))}</td><td>{_esc(mv.get("read",""))}</td></tr>')
        parts.append('</table></div>')
    if ins.get("competitor_moves") or ins.get("attack_points") or ins.get("learn_from"):
        parts.append('<div class="sec"><h2>🕵️ 竞手动态与机会</h2>')
        if ins.get("competitor_moves"):
            parts.append('<h3>本周竞手动作</h3><table><tr><th style="width:18%">品牌/ASIN</th>'
                         '<th style="width:34%">变更</th><th>解读</th></tr>')
            for mv in ins["competitor_moves"]:
                who = mv.get("brand") or mv.get("asin") or "?"
                parts.append(f'<tr><td>{_esc(who)}<br><span style="color:#999;font-size:11px">'
                             f'{_esc(mv.get("asin",""))}</span></td>'
                             f'<td>{_esc(mv.get("move",""))}</td><td>{_esc(mv.get("read",""))}</td></tr>')
            parts.append('</table>')
        if ins.get("attack_points"):
            parts.append('<h3>攻击点（竞手漏洞 × 我方覆盖）</h3><table><tr><th style="width:30%">高频问题</th>'
                         '<th style="width:22%">未覆盖的竞手</th><th>怎么打</th></tr>')
            for ap in ins["attack_points"]:
                parts.append(f'<tr><td>{_esc(ap.get("question",""))}</td>'
                             f'<td>{_esc(" / ".join(ap.get("weak_competitors", [])))}</td>'
                             f'<td>{_esc(ap.get("how_to_use",""))}</td></tr>')
            parts.append('</table>')
        if ins.get("learn_from"):
            parts.append('<h3>可借鉴</h3><table><tr><th style="width:18%">品牌</th>'
                         '<th style="width:40%">做得好的点</th><th>怎么用</th></tr>')
            for lf in ins["learn_from"]:
                parts.append(f'<tr><td>{_esc(lf.get("brand",""))}</td>'
                             f'<td>{_esc(lf.get("what",""))}</td><td>{_esc(lf.get("apply",""))}</td></tr>')
            parts.append('</table>')
        parts.append('</div>')

    # ── 赞助位哨兵 ──
    sw = current.get("sponsored_watch", {})
    sp_new_brands = wow.get("sponsored_new_brands", []) if wow else []
    if sw.get("hits") or sp_new_brands:
        SP_CLS_LABEL = {"self": "自投（常态）", "cross": "⚠️ 跨页投放", "invaded": "🚨 入侵我方页",
                        "own_enrolled": "ℹ️ 我方已enroll", "unknown": "宿主品牌未知"}
        parts.append('<div class="sec"><h2>🪧 赞助位哨兵（Sponsored Prompts）</h2>')
        parts.append('<div class="legend">挂件里的广告位（SP/SB campaign 自动延伸）。自投=品牌挂在自己 listing（常态）；'
                     '跨页/入侵我方页=投放模式变化，需要行动。单周数量波动是广告轮播抽样特性，不做周环比解读</div>')
        sp_alerts = []
        for h in sw.get("hits", []):
            if h["cls"] == "invaded":
                sp_alerts.append(f'🚨 <b>P0</b> 竞品赞助位出现在我方 {_esc(h["asin"])} 页面：「{_esc(h["prompt"])}」')
            elif h["cls"] == "cross":
                sp_alerts.append(f'⚠️ 跨页投放：{_esc(h.get("brand") or "?")} 的赞助位出现在 '
                                 f'{_esc(h.get("host_brand") or "?")}（{_esc(h["asin"])}）页面')
            elif h["cls"] == "own_enrolled":
                sp_alerts.append(f'ℹ️ 我方已被自动 enroll：{_esc(h["asin"])}「{_esc(h["prompt"])}」'
                                 f'——去广告后台 Prompts tab 核对扣费')
        if sp_new_brands:
            sp_alerts.append('🆕 品牌首次进场：' + "、".join(_esc(b) for b in sp_new_brands))
        if sp_alerts:
            parts.append('<div class="wow-box">' + "<br>".join(sp_alerts) + '</div>')
        if sw.get("hits"):
            parts.append('<table><tr><th style="width:14%">宿主ASIN</th><th style="width:14%">宿主品牌</th>'
                         '<th>赞助问题</th><th style="width:16%">判定</th></tr>')
            for h in sw["hits"]:
                parts.append(f'<tr><td>{_esc(h["asin"])}</td><td>{_esc(h.get("host_brand") or "—")}</td>'
                             f'<td>{_esc(h["prompt"])}</td><td>{SP_CLS_LABEL.get(h["cls"], h["cls"])}</td></tr>')
            parts.append('</table>')
        parts.append('</div>')

    # ── 自有 ASIN ──
    qa_lookup = _build_qa_lookup(current)
    parts.append('<div class="sec"><h2>🏠 自有 ASIN 状态</h2>')
    parts.append('<div class="legend">「未命中」= 该问题在竞品挂件高频出现、但本品挂件未展示；'
                 '不代表 Alexa 答不出（对照下方质检结果，✓质检A 表示能明确作答，仅缺曝光位）</div>')
    for own_asin, qs in own.items():
        status = own_statuses.get(own_asin, "unknown")
        parts.append(f'<span class="asin-hd">{_esc(own_asin)}</span> '
                     f'<span style="font-size:12px;color:#666">状态: {_esc(status)}</span>')
        # 在哪些关键词有自然位
        ranks = roster.get(own_asin, {}).get("keyword_ranks", {})
        if ranks:
            rank_str = " | ".join(f"{kw[:25]}→#{r}" for kw, r in ranks.items())
            parts.append(f'<div style="font-size:12px;color:#3949ab;margin:4px 0">自然位: {_esc(rank_str)}</div>')
        else:
            parts.append('<div style="font-size:12px;color:#999;margin:4px 0">未进入各关键词 Top5 自然位</div>')
        if qs:
            for q in qs:
                parts.append(f'<div class="qi">{_esc(q)}</div>')
        else:
            parts.append('<div style="font-size:12px;color:#999;margin:4px 0">暂无 Alexa 问题（新品期正常）</div>')
        # Gap（来自关键词层）
        if not qs:
            # 新品期挂件未生成：所有高频问题都会被机械列为 gap，属结构性噪音，降级为一行灰字
            gap_kw_cnt = sum(1 for kw_item in ka if kw_item["own_gap"].get(own_asin))
            if gap_kw_cnt:
                parts.append(f'<div style="font-size:12px;color:#999;margin:4px 0">'
                             f'新品期挂件未生成，{gap_kw_cnt} 个关键词的 gap 暂不计（Top5 高频问题见关键词板块）</div>')
        else:
            for kw_item in ka:
                gap = kw_item["own_gap"].get(own_asin, [])
                if not gap:
                    continue
                annotated, all_ok = [], True
                for g in gap:
                    grade = _gap_grade(qa_lookup, own_asin, g)
                    if grade == "A":
                        annotated.append(f'{_esc(g)} <span class="wn">✓质检A</span>')
                    elif grade in ("C", "D"):
                        annotated.append(f'{_esc(g)} <span class="wd">✗质检{grade}，需补内容</span>')
                        all_ok = False
                    else:
                        annotated.append(_esc(g))
                        all_ok = False
                box_cls = "gap-ok" if all_ok else "gap-box"
                icon = "✓" if all_ok else "⚡"
                parts.append(f'<div class="{box_cls}">{icon} <b>「{_esc(kw_item["keyword"])}」Top5 高频但自有未命中：</b><br>'
                              + " / ".join(annotated) + '</div>')
    parts.append('</div>')

    # ── 本品问答质检（Phase 5）──
    own_qa = current.get("own_qa", {})
    if own_qa:
        GRADE_LABEL = {"A": ("A 明确", "tag-m"), "B": ("B 含糊", "tag-h"),
                       "C": ("C 答不出", "tag-u"), "D": ("D 答错", "tag-u"), "E": ("E 未获取", "tag-u")}
        parts.append('<div class="sec"><h2>🧪 本品 Alexa 问答质检</h2>')
        parts.append('<div class="legend">分级：A明确 / B含糊 / C答不出 / D答错 / E未获取（C→A 为改善）；'
                     '标记列为脚本自动风险检测（答错·引用差评·竞品引流chip·属性缺失），— 为正常无风险</div>')
        qa_wow_list = wow.get("qa_wow", []) if wow else []
        if qa_wow_list:
            changes = "<br>".join(
                f'<b>{_esc(x["asin"])}</b> 「{_esc(x["question"])}」 {_esc(x["prev_grade"])} → <b>{_esc(x["curr_grade"])}</b>'
                for x in qa_wow_list)
            parts.append(f'<div class="wow-box">⚠️ <b>分级变化（对比上周）：</b><br>{changes}</div>')
        for own_asin, item in own_qa.items():
            parts.append(f'<span class="asin-hd">{_esc(own_asin)}</span> '
                         f'<span style="font-size:12px;color:#666">状态: {_esc(item["status"])}</span>')
            if item["results"]:
                parts.append('<table><tr><th style="width:26%">探针问题</th><th style="width:8%">分级</th>'
                             '<th>Alexa 回答</th><th style="width:22%">标记</th></tr>')
                for r in item["results"]:
                    label, cls = GRADE_LABEL.get(r["grade"], (r["grade"], "tag-h"))
                    flag_str = "<br>".join(_esc(f) for f in r["flags"]) if r["flags"] else "—"
                    parts.append(f'<tr><td>{_esc(r["question"])}</td>'
                                 f'<td><span class="{cls}">{_esc(label)}</span></td>'
                                 f'<td>{_esc(r["answer"][:300])}</td><td>{flag_str}</td></tr>')
                parts.append('</table>')
            else:
                parts.append('<div style="font-size:12px;color:#999;margin:4px 0">本次未获取到问答结果</div>')
        parts.append('</div>')

    # ── 关键词层 ──
    parts.append('<div class="sec"><h2>🔑 关键词视角 Alexa 问题</h2>')
    GROUP_PREV = ""
    for kw_item in ka:
        if kw_item["group"] != GROUP_PREV:
            parts.append(f'<h3>━ {_esc(kw_item["group"])} 品类</h3>')
            GROUP_PREV = kw_item["group"]
        # WoW 搜索量
        kw_wow_item = next((x for x in wow.get("kw_wow", []) if x["keyword"] == kw_item["keyword"]), None)
        if kw_wow_item and kw_wow_item.get("has_prev"):
            wd = kw_wow_item["weekly_delta"]
            md = kw_wow_item["monthly_delta"]
            cd = kw_wow_item["cpc_delta"]
            def _d(v): return (f'<span class="wn">↑{abs(v):,}</span>' if v>0
                               else (f'<span class="wd">↓{abs(v):,}</span>' if v<0 else '持平'))
            meta_str = (f"周搜：<b>{kw_item['weekly_search']:,}</b>（{_d(wd)}）"
                       f" | 月搜：<b>{kw_item['monthly_search']:,}</b>（{_d(md)}）"
                       f" | CPC：<b>${kw_item['cpc']:.2f}</b>（{'↑' if cd>0 else '↓' if cd<0 else ''}${abs(cd):.2f}）")
        else:
            meta_str = (f"周搜：<b>{kw_item['weekly_search']:,}</b>"
                       f" | 月搜：<b>{kw_item['monthly_search']:,}</b>"
                       f" | CPC：<b>${kw_item['cpc']:.2f}</b>")

        comp_str = (f' | 竞品数：{_esc(kw_item["competition"])}'
                    if kw_item["competition"] and kw_item["competition"] not in ("-", "None") else "")
        parts.append(f'<div class="kw-card">'
                     f'<div class="kw-title">🔑 {_esc(kw_item["keyword"])}</div>'
                     f'<div class="kw-meta">{meta_str}{comp_str}'
                     f' | Top5成功抓取：{kw_item["top5_ok_count"]}/{len(kw_item["top5"])}</div>')

        hf = kw_item["analysis"].get("high_freq", [])
        if hf:
            parts.append('<b style="font-size:12px">Top5 高频 Alexa 问题：</b>'
                         '<table style="margin-top:6px"><tr><th>问题</th><th>出现/Top5</th></tr>')
            for q, asins in hf:
                parts.append(f'<tr><td>{_esc(q)}</td><td>{len(asins)}/{kw_item["top5_ok_count"]}</td></tr>')
            parts.append('</table>')
        else:
            parts.append('<div style="font-size:12px;color:#999;margin-top:6px">Top5 暂无高频共性问题</div>')

        # Top5 ASIN 列表
        parts.append('<details style="margin-top:8px"><summary>Top5 ASIN 详情</summary><div style="padding:6px 0">')
        for entry in kw_item["top5"]:
            asin = entry["asin"]
            qs_here = current["all_questions"].get(asin, [])
            is_own = roster.get(asin, {}).get("is_own", False)
            label = " 🏠自有" if is_own else ""
            parts.append(f'<span class="asin-hd">#{entry["rank"]} {_esc(asin)}{_esc(label)}</span>'
                         f' <span style="font-size:11px;color:#666">{_esc(entry["brand"])} | ${entry["price"]:.2f} | 月销{entry["monthly_sales"]:,}</span>')
            if qs_here:
                for q in qs_here:
                    parts.append(f'<div class="qi">{_esc(q)}</div>')
            else:
                st = current["run_statuses"].get(asin, {}).get("status", "?")
                parts.append(f'<div class="qi" style="color:#999">无Alexa问题（{_esc(st)}）</div>')
        parts.append('</div></details></div>')
    parts.append('</div>')

    # ── 竞品全景（固定40个）──
    parts.append('<div class="sec"><h2>📦 竞品全景（固定竞品高频汇总）</h2>')
    if wow:
        new_q  = wow.get("comp_hf_new", [])
        drop_q = wow.get("comp_hf_drop", [])
        if new_q or drop_q:
            parts.append('<div class="wow-box">')
            if new_q:
                parts.append(f'<div class="wn">🆕 新进入高频（{len(new_q)}条）：' +
                              " / ".join(_esc(q) for q in new_q[:5]) + ('...' if len(new_q)>5 else '') + '</div>')
            if drop_q:
                parts.append(f'<div class="wd">📉 退出高频（{len(drop_q)}条）：' +
                              " / ".join(_esc(q) for q in drop_q[:5]) + '</div>')
            parts.append('</div>')

    total = ca.get("total", 1)
    threshold = ca.get("threshold", 2)
    if ca.get("high_freq"):
        parts.append(f'<h3>高频共性问题（≥{threshold} 个竞品，{threshold/total:.0%}+）</h3>'
                     '<table><tr><th>问题</th><th>竞品数</th><th>覆盖率</th></tr>')
        for q, asins in ca["high_freq"]:
            pct = len(asins)/total if total else 0
            bar_w = int(pct*100)
            parts.append(f'<tr><td>{_esc(q)}</td><td>{len(asins)}/{total}</td>'
                         f'<td><span class="bar-wrap"><span class="bar" style="width:{bar_w}%"></span></span> {pct:.0%}</td></tr>')
        parts.append('</table>')

    if ca.get("mid_freq"):
        parts.append(f'<h3>中频问题（2~{threshold-1} 个竞品）</h3>'
                     '<table><tr><th>问题</th><th>竞品数</th></tr>')
        for q, asins in ca["mid_freq"]:
            parts.append(f'<tr><td>{_esc(q)}</td><td>{len(asins)}</td></tr>')
        parts.append('</table>')
    parts.append('</div>')

    # ── 新问题/消失问题 WoW ──
    if wow and (wow.get("brand_new_q") or wow.get("disappeared_q")):
        parts.append('<div class="sec"><h2>🔄 全量问题 WoW 变化</h2>')
        if wow.get("brand_new_q"):
            parts.append(f'<details><summary class="wn">✨ 全新问题（{len(wow["brand_new_q"])}条，上周未见）</summary><div>')
            for q in wow["brand_new_q"]:
                parts.append(f'<div class="qi wn">+ {_esc(q)}</div>')
            parts.append('</div></details>')
        if wow.get("disappeared_q"):
            parts.append(f'<details><summary class="wd">👻 消失的问题（{len(wow["disappeared_q"])}条）</summary><div>')
            for q in wow["disappeared_q"]:
                parts.append(f'<div class="qi wd">- {_esc(q)}</div>')
            parts.append('</div></details>')
        parts.append('</div>')

    parts.append('</div></body></html>')
    html = "\n".join(parts)
    OUTPUT_HTML.write_text(html, encoding="utf-8")
    print(f"✅ HTML 报告已保存: {OUTPUT_HTML}")
    return OUTPUT_HTML


def push_to_github_pages(html_path: Path) -> str:
    dest = GITHUB_REPO / html_path.name
    try:
        shutil.copy2(str(html_path), str(dest))
        subprocess.run(["git", "add", html_path.name], cwd=str(GITHUB_REPO),
                       check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", f"alexa intel: {RUN_TIMESTAMP}"],
                       cwd=str(GITHUB_REPO), check=True, capture_output=True)
        subprocess.run(["git", "push"], cwd=str(GITHUB_REPO),
                       check=True, capture_output=True)
        url = f"{GITHUB_PAGES_BASE}/{html_path.name}"
        print(f"🚀 GitHub Pages: {url}")
        return url
    except Exception as e:
        print(f"❌ GitHub Pages 推送失败: {e}")
        return ""


def save_excel_report(current: dict) -> Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)

    def _style(ws):
        for c in ws[1]:
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center")

    # Sheet1: 关键词高频 Alexa 问题
    ws1 = wb.active; ws1.title = "关键词Alexa高频"
    ws1.append(["关键词", "品类", "Alexa问题", "Top5出现次数", "周搜索量"])
    _style(ws1)
    for kw in current["keyword_analysis"]:
        for q, asins in kw["analysis"].get("high_freq", []):
            ws1.append([kw["keyword"], kw["group"], q, len(asins), kw["weekly_search"]])
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["C"].width = 65

    # Sheet2: 竞品高频
    ws2 = wb.create_sheet("竞品高频问题")
    ca = current["comp_analysis"]
    total = ca.get("total", 1)
    ws2.append(["问题", "竞品数", "覆盖率"])
    _style(ws2)
    for q, asins in ca.get("high_freq", []):
        ws2.append([q, len(asins), f"{len(asins)/total:.0%}"])
    ws2.column_dimensions["A"].width = 65

    # Sheet3: 自有ASIN gap分析
    ws3 = wb.create_sheet("自有ASIN Gap分析")
    ws3.append(["自有ASIN", "关键词", "缺失高频问题"])
    _style(ws3)
    for kw in current["keyword_analysis"]:
        for own_asin, gaps in kw["own_gap"].items():
            for g in gaps:
                ws3.append([own_asin, kw["keyword"], g])
    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 65

    # Sheet4: 所有ASIN问题全览
    ws4 = wb.create_sheet("全部ASIN问题")
    ws4.append(["ASIN", "标签", "Alexa问题"])
    _style(ws4)
    roster = current["roster"]
    for asin, qs in current["all_questions"].items():
        meta = roster.get(asin, {})
        tag = "自有" if meta.get("is_own") else ("固定竞品" if meta.get("is_fixed") else "关键词新")
        for q in qs:
            ws4.append([asin, tag, q])
    ws4.column_dimensions["A"].width = 15
    ws4.column_dimensions["C"].width = 65

    # Sheet5: 本品问答质检（Phase 5）
    own_qa = current.get("own_qa", {})
    if own_qa:
        ws5 = wb.create_sheet("本品问答质检")
        ws5.append(["自有ASIN", "探针问题", "预期极性", "分级", "回答极性", "标记", "Alexa回答", "建议chips"])
        _style(ws5)
        for asin, item in own_qa.items():
            for r in item["results"]:
                ws5.append([asin, r["question"], r.get("expect") or "", r["grade"],
                            r.get("polarity") or "", "；".join(r["flags"]),
                            r["answer"], " | ".join(r["chips"])])
        ws5.column_dimensions["A"].width = 15
        ws5.column_dimensions["B"].width = 40
        ws5.column_dimensions["F"].width = 40
        ws5.column_dimensions["G"].width = 80

    wb.save(str(OUTPUT_EXCEL))
    print(f"✅ Excel 已保存: {OUTPUT_EXCEL}")
    return OUTPUT_EXCEL


def send_feishu(current: dict, wow: dict, html_url: str):
    ka = current["keyword_analysis"]
    ca = current["comp_analysis"]
    own = current["own_q"]
    own_statuses = current["own_status"]
    roster = current["roster"]
    total_scraped = len(current["all_questions"])
    ok_count = sum(1 for r in current["run_statuses"].values() if r["status"] == "ok")

    elements = [{
        "tag": "div",
        "text": {"tag": "lark_md",
                 "content": f"**生成时间：** {RUN_DATETIME}　**抓取成功：** {ok_count}/{total_scraped}"}
    }, {"tag": "hr"}]

    # ── 模块零：洞察置顶（决策先行）───────────────────────────
    ins = current.get("insights", {}) or {}
    if ins.get("summary_line") or ins.get("actions"):
        top_lines = []
        if ins.get("summary_line"):
            top_lines.append(f"**📌 {ins['summary_line']}**")
        for a in ins.get("actions", [])[:3]:
            top_lines.append(f"**{a.get('priority','')}**［{a.get('target','')}］{a.get('action','')}")
        n_more = len(ins.get("actions", [])) - 3
        if n_more > 0:
            top_lines.append(f"_…另 {n_more} 条行动建议见 HTML_")
        elements.append({"tag": "div", "text": {"tag": "lark_md",
                         "content": "**🎯 本周行动建议**\n" + "\n".join(top_lines)}})
        elements.append({"tag": "hr"})
    if ins.get("own_listing_moves"):
        ol_lines = ["**🏠 本品 Listing 本周变化**"]
        for mv in ins["own_listing_moves"][:5]:
            ol_lines.append(f"- **{mv.get('asin','')}**：{mv.get('move','')}\n  ↳ {mv.get('read','')}")
        elements.append({"tag": "div", "text": {"tag": "lark_md", "content": "\n".join(ol_lines)}})
        elements.append({"tag": "hr"})
    if ins.get("competitor_moves") or ins.get("attack_points"):
        cm_lines = ["**🕵️ 竞手动态与机会**"]
        for mv in ins.get("competitor_moves", [])[:3]:
            who = mv.get("brand") or mv.get("asin") or "?"
            cm_lines.append(f"- **{who}**：{mv.get('move','')}\n  ↳ {mv.get('read','')}")
        for ap in ins.get("attack_points", [])[:2]:
            cm_lines.append(f"⚔️ 攻击点「{ap.get('question','')}」竞手 "
                            f"{'/'.join(ap.get('weak_competitors', [])[:3])} 未覆盖\n  ↳ {ap.get('how_to_use','')}")
        elements.append({"tag": "div", "text": {"tag": "lark_md", "content": "\n".join(cm_lines)}})
        elements.append({"tag": "hr"})

    # ── 模块一：竞品 ──────────────────────────────────────────
    elements.append({"tag": "div", "text": {"tag": "lark_md",
                     "content": "**📦 【竞品】高频 Alexa 问题**"}})

    # 竞品整体高频
    hf = ca.get("high_freq", [])
    hf_lines = "\n".join(f"{i+1}. {q}（{len(a)}/{ca.get('total', 1)} 个ASIN）"
                         for i, (q, a) in enumerate(hf[:8])) if hf else "暂无高频问题"
    elements.append({"tag": "div", "text": {"tag": "lark_md", "content": hf_lines}})
    elements.append({"tag": "hr"})

    # ── 赞助位哨兵：异常才展开，平时一行 ──
    sw = current.get("sponsored_watch", {})
    sp_new_brands = wow.get("sponsored_new_brands", []) if wow else []
    sp_alerts = []
    for h in sw.get("hits", []):
        if h["cls"] == "invaded":
            sp_alerts.append(f'🚨 **P0** 竞品赞助位入侵我方页面 {h["asin"]}：{h["prompt"]}')
        elif h["cls"] == "cross":
            sp_alerts.append(f'⚠️ 跨页投放：{h.get("brand") or "?"} → '
                             f'{h.get("host_brand") or "?"}（{h["asin"]}）')
        elif h["cls"] == "own_enrolled":
            sp_alerts.append(f'ℹ️ 我方被自动enroll（{h["asin"]}），查广告后台 Prompts tab 扣费')
    if sp_new_brands:
        sp_alerts.append("🆕 赞助位品牌首次进场：" + "、".join(sp_new_brands))
    if sp_alerts:
        elements.append({"tag": "div", "text": {"tag": "lark_md",
                         "content": "**🪧 赞助位哨兵**\n" + "\n".join(sp_alerts)}})
    else:
        n_hits = len(sw.get("hits", []))
        sp_brands = sorted({h["brand"] for h in sw.get("hits", []) if h.get("brand")})
        heartbeat = (f"🪧 赞助位哨兵：本周 {n_hits} 条自投（{'、'.join(sp_brands)}），无异常"
                     if n_hits else "🪧 赞助位哨兵：本周未见赞助位，无异常")
        elements.append({"tag": "div", "text": {"tag": "lark_md", "content": heartbeat}})
    elements.append({"tag": "hr"})

    # 各关键词 Top5 高频
    curr_group = ""
    for kw_item in ka:
        if kw_item["group"] != curr_group:
            curr_group = kw_item["group"]
            elements.append({"tag": "div", "text": {"tag": "lark_md",
                             "content": f"**━━ {curr_group} 品类 ━━**"}})
        lines = [f"**🔑 {kw_item['keyword']}**　"
                 f"周搜 {kw_item['weekly_search']:,} | 月搜 {kw_item['monthly_search']:,} | CPC ${kw_item['cpc']:.2f}"]
        hf_kw = kw_item["analysis"].get("high_freq", [])
        if hf_kw:
            lines.append(f"Top5自然位高频（{kw_item['top5_ok_count']}个ASIN）：")
            for q, asins in hf_kw[:5]:
                lines.append(f"- {q}（{len(asins)}/{kw_item['top5_ok_count']}）")
        else:
            lines.append("_Top5 暂无高频共性问题_")
        elements.append({"tag": "div", "text": {"tag": "lark_md", "content": "\n".join(lines)}})
    elements.append({"tag": "hr"})

    # ── 模块二：本品 ──────────────────────────────────────────
    elements.append({"tag": "div", "text": {"tag": "lark_md",
                     "content": "**🏠 【本品】Alexa 问题**"}})
    own_lines = []
    for own_asin, qs in sorted(own.items()):
        status = own_statuses.get(own_asin, "?")
        ranks = roster.get(own_asin, {}).get("keyword_ranks", {})
        rank_str = " | ".join(f"{k[:20]}→#{v}" for k, v in ranks.items()) if ranks else "未进入Top5自然位"
        own_lines.append(f"**{own_asin}**　抓取状态：{status}　自然位：{rank_str}")
        if qs:
            own_lines.append(f"当周被问（{len(qs)} 条）：")
            for q in qs:
                own_lines.append(f"- {q}")
        else:
            own_lines.append("当周暂无 Alexa 问题（新品期正常）")
        if not qs:
            # 新品期挂件未生成：gap 全是结构性噪音，不逐条列
            gap_kw_cnt = sum(1 for kw_item in ka if kw_item["own_gap"].get(own_asin))
            if gap_kw_cnt:
                own_lines.append(f"Gap：新品期挂件未生成，{gap_kw_cnt} 个关键词 gap 暂不计")
        else:
            qa_lookup = _build_qa_lookup(current)
            gaps_warn, gaps_ok = [], []
            for kw_item in ka:
                gap = kw_item["own_gap"].get(own_asin, [])
                if not gap:
                    continue
                annotated, has_bad = [], False
                for g in gap[:3]:
                    grade = _gap_grade(qa_lookup, own_asin, g)
                    if grade == "A":
                        annotated.append(f"{g} ✓A")
                    elif grade in ("C", "D"):
                        annotated.append(f"{g} ✗{grade}")
                        has_bad = True
                    else:
                        annotated.append(g)
                        has_bad = True
                line = f"「{kw_item['keyword']}」" + " / ".join(annotated)
                (gaps_warn if has_bad else gaps_ok).append(line)
            if gaps_warn:
                own_lines.append("⚡ Gap 需补内容（✗=质检答不出，无标=未探针）：")
                own_lines.extend(gaps_warn)
            if gaps_ok:
                own_lines.append("✅ Gap 已覆盖（质检A，仅挂件未展示）：")
                own_lines.extend(gaps_ok)
    if not own_lines:
        own_lines.append("未配置自有 ASIN")
    elements.append({"tag": "div", "text": {"tag": "lark_md", "content": "\n".join(own_lines)}})

    # ── 模块三：本品问答质检（Phase 5）─────────────────────────
    # 全绿时压缩为一行；有 B/C/D/E 或分级变化才展开明细
    own_qa = current.get("own_qa", {})
    if own_qa:
        elements.append({"tag": "hr"})
        GRADE_EMOJI = {"A": "🟢A", "B": "🟡B", "C": "🟠C", "D": "🔴D", "E": "⚪E"}
        qa_wow_list = wow.get("qa_wow", []) if wow else []
        all_results = [(a, r) for a, item in sorted(own_qa.items()) for r in item["results"]]
        bad = [(a, r) for a, r in all_results
               if r["grade"] != "A" or any(f.startswith(("⚠️", "🔀")) or "相反" in f for f in r["flags"])]
        qa_lines = ["**🧪 【质检】Alexa 回答本品问题的能力**"]
        if not bad and not qa_wow_list:
            qa_lines.append(f"🟢 全部通过：{len(all_results)} 个探针问题全 A（明确作答），无引流/差评/答错标记")
        else:
            n_ok = len(all_results) - len(bad)
            if bad:
                qa_lines.append(f"🟢 {n_ok} 项通过；需要关注 {len(bad)} 项：")
            else:
                qa_lines.append(f"🟢 全部 {n_ok} 项通过；本周分级有变化：")
            for a, r in bad:
                mark = GRADE_EMOJI.get(r["grade"], r["grade"])
                alert = "　" + "；".join(f for f in r["flags"] if f.startswith(("⚠️", "🔀")) or "相反" in f) \
                        if r["flags"] else ""
                qa_lines.append(f"{mark} [{a}] {r['question']}{alert}")
            for x in qa_wow_list:
                qa_lines.append(f"⚠️ 分级变化 {x['asin']}「{x['question'][:30]}」{x['prev_grade']}→{x['curr_grade']}")
            qa_lines.append("_A明确 B含糊 C答不出 D答错 E未获取；回答原文见 HTML_")
        elements.append({"tag": "div", "text": {"tag": "lark_md", "content": "\n".join(qa_lines)}})

    if html_url:
        elements.append({"tag": "action", "actions": [{
            "tag": "button",
            "text": {"tag": "plain_text", "content": "📄 查看完整 HTML 报告"},
            "type": "primary",
            "url": html_url}]})
    elements.append({"tag": "note", "elements": [{"tag": "plain_text",
                     "content": f"Alexa 综合情报 | {RUN_TIMESTAMP} | 关键词Top{KEYWORD_TOP_N}+固定竞品+自有"}]})

    payload = {
        "msg_type": "interactive",
        "card": {
            "header": {"title": {"tag": "plain_text", "content": "📊 Alexa 问题综合情报周报"},
                       "template": "indigo"},
            "elements": elements,
        }
    }
    try:
        resp = requests.post(FEISHU_WEBHOOK, headers={"Content-Type": "application/json"},
                             data=json.dumps(payload, ensure_ascii=False), timeout=10)
        resp.raise_for_status()
        r = resp.json()
        if r.get("code") == 0 or r.get("StatusCode") == 0:
            print("✅ 飞书已发送")
        else:
            print(f"⚠ 飞书返回: {r}")
    except Exception as e:
        print(f"✗ 飞书失败: {e}")


# ─── 主流程入口（分步测试时可从任一阶段启动）─────────────────────────────────

def main():
    print(f"═══ Alexa 综合情报 {RUN_DATETIME} ═══════════════════════════")

    # Phase 1: Sorftime 关键词查询 → Top5 ASIN
    keyword_data = run_sorftime_phase()

    # Phase 2: 构建统一 ASIN 名单
    roster = build_asin_roster(keyword_data)
    own_count = sum(1 for m in roster.values() if m["is_own"])
    fix_count = sum(1 for m in roster.values() if m["is_fixed"] and not m["is_own"])
    new_count = sum(1 for m in roster.values() if not m["is_fixed"] and not m["is_own"])
    print(f"\n✅ 名单汇总：自有={own_count} 固定竞品={fix_count} 关键词新={new_count} 合计={len(roster)}")

    # Phase 3: Playwright 抓取
    asin_questions, run_statuses = run_playwright_phase(roster)

    # Phase 4: 三层分析
    current = run_analysis_phase(keyword_data, roster, asin_questions, run_statuses)

    # Phase 5: 本品 Alexa 问答质检
    own_set = {a for a, m in roster.items() if m["is_own"]}
    current["own_qa"] = run_alexa_qa_phase(own_set)

    # WoW：必须先读上周存档再覆盖，否则永远在和本次自比
    prev = load_prev_archive()
    wow  = compute_wow(current, prev) if prev else {}
    save_archive(current)

    # Phase 6: 洞察合成（两个目标：本品 Alexa 推荐优化 / 竞手动态与机会）
    current["insights"] = run_insight_phase(build_insight_facts(current, wow, prev))

    # Phase 7: 输出
    html_path = generate_html(current, wow)
    html_url  = push_to_github_pages(html_path)
    save_excel_report(current)
    send_feishu(current, wow, html_url)

    print(f"\n✅ 全部完成 — {RUN_DATETIME}")


def main_qa_only():
    """只跑 Phase 5 问答质检（调试/手动验证用），不发飞书、不写存档。"""
    print(f"═══ Phase 5 单独测试 {RUN_DATETIME} ═══════════════════════════")
    own_asins = set(_load_asin_file(OWN_ASINS_FILE))
    qa = run_alexa_qa_phase(own_asins)
    out = RAW_DIR / f"alexa_qa_test_{RUN_TIMESTAMP}.json"
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n✅ 测试结果已保存: {out}")
    for asin, item in qa.items():
        print(f"\n{asin} [{item['status']}]")
        for r in item["results"]:
            print(f"  [{r['grade']}] {r['question']}")
            print(f"      答: {r['answer'][:150]}")
            if r["flags"]:
                print(f"      标记: {'；'.join(r['flags'])}")


if __name__ == "__main__":
    if "--login-setup" in sys.argv:
        main_login_setup()
    elif "--qa-only" in sys.argv:
        main_qa_only()
    else:
        main()
