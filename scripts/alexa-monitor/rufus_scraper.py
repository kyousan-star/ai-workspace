"""
rufus_scraper.py
抓取 Amazon listing 详情页上的 Alexa for Shopping / Rufus 问题，分析高频/共性/差异性问题，并推送结果到飞书群。
"""

import json
import math
import os
import re
import time
import unicodedata
from pathlib import Path
from datetime import datetime
from collections import Counter, defaultdict

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.sync_api import sync_playwright

# ─── 配置 ────────────────────────────────────────────────────────────────────

OUTPUT_DIR = Path(__file__).parent
REPORT_DIR = OUTPUT_DIR / "reports"
RAW_DIR = OUTPUT_DIR / "raw"
DEBUG_DIR = OUTPUT_DIR / "debug"
ASINS_FILE = OUTPUT_DIR / "asins.txt"
ENV_FILE = OUTPUT_DIR / ".env"

DEFAULT_ASINS = [
    "B0BPM21DF1",
    "B0DC3RW27T",
    "B0DTHQKCRK",
    "B0CRDP2QH9",
    "B0DHTZJ23J",
    "B0DLW5BMYF",
    "B0DVLXMTQG",
    "B0CRYTX9YX",
    "B0CZDY4Q28",
    "B0BDZ27CGQ",
    "B0CQJVXT74",
    "B0FB94JGT7",
    "B0F7XDXD69",
    "B0BCG1V4RN",
    "B0DJT2YGK8",
    "B0DPYNH1DZ",
    "B0DNQC7X4K",
    "B0CQP9S8Q2",
    "B0BNTCKQRF",
    "B0DXFJBX24",
]

DEFAULT_WEBHOOK_URL = "https://open.feishu.cn/open-apis/bot/v2/hook/20e06d51-6ac0-4e78-8229-0dd3abd581b3"

DEFAULT_GDRIVE_FOLDER_ID = "1YCHa8JCIZjv_8FRm_QVDp2FIFUocEwtJ"
DEFAULT_GDRIVE_TOKEN = Path("/Users/lihuan/.claude/scripts/gdrive_token.json")

DEFAULT_COOKIES_FILE = OUTPUT_DIR / "amazon_cookies.json"

DELAY_BETWEEN_ASINS = int(os.getenv("RUFUS_DELAY_BETWEEN_ASINS", "10"))  # 秒，避免触发反爬
MAX_RETRIES = int(os.getenv("RUFUS_MAX_RETRIES", "3"))
SECOND_PASS_RETRIES = int(os.getenv("RUFUS_SECOND_PASS_RETRIES", "1"))
NAVIGATION_TIMEOUT_MS = int(os.getenv("RUFUS_NAVIGATION_TIMEOUT_MS", "60000"))
DOM_READY_TIMEOUT_MS = int(os.getenv("RUFUS_DOM_READY_TIMEOUT_MS", "15000"))
RUFUS_HINT_TIMEOUT_MS = int(os.getenv("RUFUS_HINT_TIMEOUT_MS", "20000"))
HEADLESS = os.getenv("RUFUS_HEADLESS", "1").lower() not in {"0", "false", "no"}
AI_ASSISTANT_LABEL = "Alexa for Shopping"
AI_ASSISTANT_SHORT = "Alexa"
MIN_SUCCESSFUL_ASINS = int(os.getenv("AI_ASSISTANT_MIN_SUCCESSFUL_ASINS", "15"))
MAX_NOT_FOUND_ASINS = int(os.getenv("AI_ASSISTANT_MAX_NOT_FOUND_ASINS", "5"))
TRANSIENT_STATUSES = {"timeout", "error"}
BLOCKED_RESOURCE_TYPES = {"image", "media", "font"}

# 运行时间戳（整个脚本只生成一次，MD 和 Excel 用同一个）
RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d")
RUN_DATETIME = datetime.now().strftime("%Y-%m-%d %H:%M")

OUTPUT_EXCEL = RAW_DIR / f"rufus_questions_{RUN_TIMESTAMP}.xlsx"
OUTPUT_MD = REPORT_DIR / f"rufus_analysis_report_{RUN_TIMESTAMP}.md"


# ─── 工具函数 ──────────────────────────────────────────────────────────────────

def load_env_file(path: Path = ENV_FILE):
    """加载简单 KEY=VALUE 格式的本地配置，不覆盖系统环境变量。"""
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def load_asins(path: Path = ASINS_FILE) -> list:
    """优先从 asins.txt 读取 ASIN；不存在时使用脚本内默认清单。"""
    if not path.exists():
        return DEFAULT_ASINS

    asins = []
    seen = set()
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.split("#", 1)[0].strip()
        if not line:
            continue
        asin = re.sub(r"[^A-Z0-9]", "", line.upper())
        if asin and asin not in seen:
            asins.append(asin)
            seen.add(asin)
    return asins or DEFAULT_ASINS


load_env_file()
ASINS = load_asins()
WEBHOOK_URL = os.getenv("RUFUS_FEISHU_WEBHOOK_URL", DEFAULT_WEBHOOK_URL)
GDRIVE_FOLDER_ID = os.getenv("RUFUS_GDRIVE_FOLDER_ID", DEFAULT_GDRIVE_FOLDER_ID)
GDRIVE_TOKEN = Path(os.getenv("RUFUS_GDRIVE_TOKEN", str(DEFAULT_GDRIVE_TOKEN))).expanduser()
COOKIES_FILE = Path(os.getenv("RUFUS_COOKIES_FILE", str(DEFAULT_COOKIES_FILE))).expanduser()


def normalize_question(text: str) -> str:
    """清除所有隐藏/零宽字符，统一 Unicode，折叠多余空格"""
    # 替换各类不可见字符为普通空格
    text = re.sub(r'[\xa0\u200b\u200c\u200d\u200e\u200f\u00ad\ufeff\u2060]', ' ', text)
    # Unicode NFC 归一化（消除同形异码）
    text = unicodedata.normalize('NFC', text)
    # 折叠多余空格
    text = re.sub(r' +', ' ', text)
    return text.strip()


IGNORED_QUESTION_PATTERNS = [
    re.compile(r"^would you like to tell us about a lower price\?$", re.I),
    re.compile(r"^what do customers say\?$", re.I),
    re.compile(r"^what types of things can i ask\?$", re.I),
]


def filter_questions(texts: list) -> list:
    """只保留自然语言问题，并排除 Amazon 页面通用控件问题。"""
    questions = []
    for text in texts:
        question = normalize_question(text)
        if not _is_real_question(question):
            continue
        if any(pattern.search(question) for pattern in IGNORED_QUESTION_PATTERNS):
            continue
        questions.append(question)
    return list(dict.fromkeys(questions))


# ─── 抓取单个 ASIN ──────────────────────────────────────────────────────────────

def scrape_rufus_questions(asin: str, context) -> dict:
    """抓取单个 ASIN，返回 status + questions，便于后续统计失败原因。"""
    last_result = {"asin": asin, "status": "unknown", "questions": [], "error": ""}
    for attempt in range(1, MAX_RETRIES + 1):
        if attempt > 1:
            print(f"    ↻ 重试第 {attempt}/{MAX_RETRIES} 次")
            time.sleep(min(10, 2 + attempt * 2))
        last_result = _scrape_rufus_questions_once(asin, context)
        if last_result["status"] in {"ok", "blocked", "captcha", "not_found"}:
            break
    return last_result


def _goto_product_page(page, url: str):
    """Amazon DP 页面经常被慢资源拖住，先建立页面再让后续 DOM 策略接管。"""
    try:
        page.goto(url, timeout=NAVIGATION_TIMEOUT_MS, wait_until="commit")
    except Exception as e:
        message = str(e)
        if "Timeout" not in message and "timeout" not in message.lower():
            raise
        print(f"    ⚠ 页面连接阶段超时，尝试继续读取已加载内容: {e}")

    try:
        page.wait_for_load_state("domcontentloaded", timeout=DOM_READY_TIMEOUT_MS)
    except Exception:
        print("    ⚠ DOMContentLoaded 等待超时，继续尝试解析当前页面")


def _wait_for_rufus_hint(page):
    """等待 AI 购物助手容器或常见商品内容出现，避免固定等待时间过短。"""
    selector = (
        "#dpx-nice-widget-container, "
        "#nile-inline-btf_feature_div, "
        "[data-feature-name='nile-inline'], "
        "[data-csa-c-slot-id*='nile'], "
        "[id*='rufus'], [class*='rufus'], "
        "[id*='alexa'], [class*='alexa'], "
        "#productTitle, #feature-bullets"
    )
    try:
        page.wait_for_selector(selector, timeout=RUFUS_HINT_TIMEOUT_MS)
    except Exception:
        print(f"    ⚠ {AI_ASSISTANT_SHORT} / 商品主体容器等待超时，继续滚动触发懒加载")


def _page_has_product_content(page) -> bool:
    """判断页面至少加载到了商品详情主体，避免空白页被误判为 no_rufus。"""
    try:
        return page.locator("#productTitle, #feature-bullets, #dp-container").count() > 0
    except Exception:
        return False


def _scrape_rufus_questions_once(asin: str, context) -> dict:
    """每次尝试使用独立 page，失败后由调用方决定是否重试。"""
    url = f"https://www.amazon.com/dp/{asin}"
    print(f"  访问: {url}")

    page = context.new_page()
    try:
        _goto_product_page(page, url)

        # 检测 503 / robot check
        title = page.title()
        content = page.content()
        content_preview = content[:500]
        content_lc = content.lower()
        if ("couldn't find that page" in content_lc
                or "could not find that page" in content_lc
                or "page not found" in title.lower()):
            print("    ⚠ Amazon 返回页面不存在/不可访问")
            _save_debug(page, asin, "not_found")
            return {"asin": asin, "status": "not_found", "questions": [], "error": "Amazon page not found"}

        if ("503" in title or "robot" in title.lower() or "captcha" in title.lower()
                or "Service Unavailable" in content_preview
                or "Type the characters" in content_preview):
            status = "captcha" if "captcha" in title.lower() or "Type the characters" in content_preview else "blocked"
            print(f"    ⚠ Amazon 封锁页（{title}），跳过")
            _save_debug(page, asin, status)
            return {"asin": asin, "status": status, "questions": [], "error": title}

        # 等待页面 JS 执行
        _wait_for_rufus_hint(page)
        page.wait_for_timeout(2500)

        # 多步滚动，触发懒加载
        for pct in [0.15, 0.3, 0.5, 0.7, 0.9, 1.0]:
            page.evaluate(f"window.scrollTo(0, document.body.scrollHeight * {pct})")
            page.wait_for_timeout(1200)

        # 回到顶部再等一次（有些 AI 购物助手入口在顶部附近）
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(2500)

        questions = _extract_questions(page)

        # 去重（保留顺序），消除 \xa0 变体造成的重复
        questions = filter_questions(questions)

        if not questions:
            if not _page_has_product_content(page):
                _save_debug(page, asin, "error")
                message = "Product page did not load enough content"
                print(f"    ✗ 抓取失败: {message}")
                return {"asin": asin, "status": "timeout", "questions": [], "error": message}
            _save_debug(page, asin, "no_rufus")
            print(f"    ⚠ 未找到 {AI_ASSISTANT_SHORT} 问题（已保存 debug HTML）")
            return {"asin": asin, "status": "no_rufus", "questions": [], "error": ""}
        else:
            print(f"    ✓ 找到 {len(questions)} 个问题")

        return {"asin": asin, "status": "ok", "questions": questions, "error": ""}

    except Exception as e:
        print(f"    ✗ 抓取失败: {e}")
        try:
            _save_debug(page, asin, "error")
        except Exception:
            pass
        status = "timeout" if "Timeout" in str(e) or "timeout" in str(e).lower() else "error"
        return {"asin": asin, "status": status, "questions": [], "error": str(e)}
    finally:
        try:
            page.close()
        except Exception:
            pass


def _extract_questions(page) -> list:
    """多策略提取 Alexa for Shopping / Rufus 问题"""

    # ── 策略 1：AI 购物助手专属容器内的按钮/chip ─────────────────────────────
    rufus_selectors = [
        "#dpx-nice-widget-container",
        "#nile-inline-btf_feature_div",
        "[data-feature-name='nile-inline']",
        "[data-feature-name*='alexa']",
        "[data-feature-name*='shopping']",
        "[data-csa-c-slot-id*='nile']",
        "[data-csa-c-slot-id*='alexa']",
        "[data-csa-c-slot-id*='shopping']",
        "#rufus-t3",
        "#ask-btf_feature_div",
        "[data-feature-name='rufus']",
        "[id*='rufus']",
        "[class*='rufus']",
        "[id*='alexa']",
        "[class*='alexa']",
        "[aria-label*='Alexa']",
        "[aria-label*='shopping']",
        "[data-csa-c-slot-id*='rufus']",
        "[data-component-type*='rufus']",
        # 新版可能的选择器
        "#aplus_feature_div [data-feature='rufus']",
        "#aplus_feature_div [data-feature*='alexa']",
        ".a-section.askATF",
        "#ask_lazy_load_div",
    ]

    for sel in rufus_selectors:
        try:
            container = page.query_selector(sel)
            if not container:
                continue
            buttons = container.query_selector_all("button, [role='button'], span.a-size-base")
            texts = [normalize_question(b.inner_text()) for b in buttons]
            texts = [t for t in texts if t and len(t) > 8 and len(t) < 200]
            if texts:
                print(f"    [策略1] 选择器 '{sel}' 找到 {len(texts)} 个")
                return texts
        except Exception:
            continue

    # ── 策略 2：全页面所有以 ? 结尾的按钮文本 ─────────────────────────────────
    try:
        results = page.evaluate("""
            () => {
                const questions = new Set();
                const candidates = document.querySelectorAll(
                    'button, [role="button"], .a-button-text, ' +
                    'span[data-action], li[data-component-type]'
                );
                candidates.forEach(el => {
                    const text = (el.innerText || el.textContent || '')
                        .replace(/\\u00a0/g, ' ').trim();
                    if (text.endsWith('?') && text.length > 8 && text.length < 200
                            && !text.includes('\\n')) {
                        questions.add(text);
                    }
                });
                return [...questions];
            }
        """)
        filtered = [q for q in results if _is_real_question(q)] if results else []
        if filtered:
            print(f"    [策略2] 全页面 ? 按钮找到 {len(filtered)} 个")
            return filtered
    except Exception:
        pass

    # ── 策略 3：更宽泛 - 任意可见文本中以 ? 结尾的短句 ───────────────────────
    try:
        results = page.evaluate("""
            () => {
                const questions = new Set();
                const walk = document.createTreeWalker(
                    document.body, NodeFilter.SHOW_TEXT, null, false
                );
                let node;
                while ((node = walk.nextNode())) {
                    const text = node.textContent.replace(/\\u00a0/g, ' ').trim();
                    if (text.endsWith('?') && text.length > 10 && text.length < 180
                            && !text.includes('\\n')) {
                        const parent = node.parentElement;
                        const ancestor = parent?.closest(
                            '[id*="rufus"], [class*="rufus"], [id*="alexa"], [class*="alexa"], ' +
                            '[id*="nile"], [class*="nile"], [id*="dpx-rex"], [class*="dpx-rex"], ' +
                            '[id*="ask"], button, [role="button"]'
                        );
                        if (ancestor) {
                            questions.add(text);
                        }
                    }
                }
                return [...questions];
            }
        """)
        filtered = [q for q in results if _is_real_question(q)] if results else []
        if filtered:
            print(f"    [策略3] 宽泛文本搜索找到 {len(filtered)} 个")
            return filtered
    except Exception:
        pass

    return []


def _is_real_question(text: str) -> bool:
    """过滤掉 JS 代码片段，只保留真实的自然语言问题"""
    t = normalize_question(text)
    if not t.endswith('?'):
        return False
    if len(t) < 10 or len(t) > 160:
        return False
    # 不能包含代码特征符号
    code_chars = set('{};=><|&+*%@#^~`$\\/')
    if any(c in t for c in code_chars):
        return False
    # 不能包含 JS 关键词
    import re
    if re.search(r'\b(var|let|const|function|return|typeof|undefined|null\b.*null\b|instanceof|prototype)\b', t):
        return False
    # 必须以大写字母开头
    if not t[0].isupper():
        return False
    # 至少有 2 个空格（即 3 个单词以上）
    if t.count(' ') < 2:
        return False
    return True


def _save_debug(page, asin: str, tag: str):
    """保存 HTML 和截图用于排查"""
    try:
        DEBUG_DIR.mkdir(exist_ok=True)
        stem = f"{asin}_{tag}"
        page.screenshot(path=str(DEBUG_DIR / f"{stem}.png"), full_page=False)
        html = page.content()
        (DEBUG_DIR / f"{stem}.html").write_text(html, encoding="utf-8")
    except Exception:
        pass


def retry_transient_failures(context, asin_questions: dict, run_statuses: dict):
    """首轮结束后补抓 timeout/error，降低 Amazon 临时加载失败对样本数的影响。"""
    retry_asins = [
        asin for asin, item in run_statuses.items()
        if item.get("status") in TRANSIENT_STATUSES
    ]
    if not retry_asins or SECOND_PASS_RETRIES <= 0:
        return

    print("\n─── 二次补抓临时失败 ASIN ───────────────────────────")
    for asin in retry_asins:
        previous_status = run_statuses[asin].get("status", "unknown")
        print(f"[补抓] ASIN: {asin}（原状态：{previous_status}）")
        best_result = run_statuses[asin]
        for attempt in range(1, SECOND_PASS_RETRIES + 1):
            if attempt > 1:
                print(f"    ↻ 补抓重试第 {attempt}/{SECOND_PASS_RETRIES} 次")
            time.sleep(DELAY_BETWEEN_ASINS)
            result = _scrape_rufus_questions_once(asin, context)
            if result["status"] == "ok":
                best_result = result
                break
            if best_result.get("status") in TRANSIENT_STATUSES and result["status"] == "no_rufus":
                best_result = result
        run_statuses[asin] = best_result
        asin_questions[asin] = best_result["questions"]


# ─── 分析 ──────────────────────────────────────────────────────────────────────

def analyze_questions(asin_questions: dict) -> dict:
    total = len(asin_questions)

    question_to_asins = defaultdict(list)
    for asin, questions in asin_questions.items():
        for q in questions:
            if asin not in question_to_asins[q]:
                question_to_asins[q].append(asin)

    # 高频：≥35% ASIN 覆盖（17个ASIN时=6个）
    threshold = max(2, math.ceil(total * 0.35))
    high_freq = sorted(
        [(q, asins) for q, asins in question_to_asins.items() if len(asins) >= threshold],
        key=lambda x: len(x[1]), reverse=True
    )

    # 中频：2~(threshold-1) 个ASIN
    mid_freq = sorted(
        [(q, asins) for q, asins in question_to_asins.items() if 2 <= len(asins) < threshold],
        key=lambda x: len(x[1]), reverse=True
    )

    unique_questions = {
        asin: [q for q in qs if len(question_to_asins[q]) == 1]
        for asin, qs in asin_questions.items()
    }

    return {
        "total_asins": total,
        "threshold": threshold,
        "question_to_asins": dict(question_to_asins),
        "high_freq": high_freq,
        "mid_freq": mid_freq,
        "unique_questions": unique_questions,
        "question_counter": Counter(q for qs in asin_questions.values() for q in qs),
    }


# ─── WoW 对比分析 ─────────────────────────────────────────────────────────────

def compare_wow(current_questions: dict, prev_questions: dict,
                current_analysis: dict, prev_scraped_at: str = "") -> dict:
    """对比本周与上周 AI 购物助手问题变化，返回结构化差异数据。"""
    if not prev_questions:
        return None

    prev_analysis = analyze_questions(prev_questions)

    prev_hf_set = {q for q, _ in prev_analysis["high_freq"]}
    curr_hf_set = {q for q, _ in current_analysis["high_freq"]}

    # 全量问题集合（去重）
    prev_all = set(q for qs in prev_questions.values() for q in qs)
    curr_all = set(q for qs in current_questions.values() for q in qs)

    # 每个ASIN级别的变化
    asin_changes = {}
    for asin in sorted(set(current_questions) | set(prev_questions)):
        curr_qs = set(current_questions.get(asin, []))
        prev_qs = set(prev_questions.get(asin, []))
        added = sorted(curr_qs - prev_qs)
        removed = sorted(prev_qs - curr_qs)
        if added or removed:
            asin_changes[asin] = {"added": added, "removed": removed}

    # 高频问题的覆盖率变化（同一问题本周vs上周ASIN数）
    hf_coverage_change = []
    all_hf_questions = curr_hf_set | prev_hf_set
    curr_q2a = current_analysis["question_to_asins"]
    prev_q2a = prev_analysis["question_to_asins"]
    total_curr = current_analysis["total_asins"]
    total_prev = prev_analysis["total_asins"]
    for q in sorted(all_hf_questions):
        curr_cnt = len(curr_q2a.get(q, []))
        prev_cnt = len(prev_q2a.get(q, []))
        if curr_cnt != prev_cnt:
            hf_coverage_change.append({
                "question": q,
                "prev": prev_cnt,
                "curr": curr_cnt,
                "total_prev": total_prev,
                "total_curr": total_curr,
            })

    return {
        "prev_scraped_at": prev_scraped_at,
        "high_freq_new": sorted(curr_hf_set - prev_hf_set),
        "high_freq_dropped": sorted(prev_hf_set - curr_hf_set),
        "brand_new_questions": sorted(curr_all - prev_all),
        "disappeared_questions": sorted(prev_all - curr_all),
        "asin_changes": asin_changes,
        "hf_coverage_change": hf_coverage_change,
        "prev_high_freq": prev_analysis["high_freq"],
        "prev_total_asins": total_prev,
    }


def latest_archive_run(archive: dict) -> dict:
    """兼容旧单次归档和新 history 归档，返回最近一次有效运行。"""
    history = archive.get("history")
    if isinstance(history, list) and history:
        return history[-1]
    return archive


def previous_archive_run(archive: dict, current_date: str) -> dict:
    """返回当前日期之前最近一次有效运行，避免同日补跑把基线变成今天。"""
    history = archive.get("history")
    candidates = history if isinstance(history, list) and history else [archive]
    for run in reversed(candidates):
        scraped_at = run.get("scraped_at", "")
        if scraped_at and not scraped_at.startswith(current_date):
            return run
    return {}


def build_archive_run(deduped_questions: dict, run_statuses: dict) -> dict:
    failed_asins = [asin for asin, item in run_statuses.items() if item.get("status") != "ok"]
    return {
        "assistant_label": AI_ASSISTANT_LABEL,
        "assistant_short": AI_ASSISTANT_SHORT,
        "source_ui_label": "alexa for shopping / Ask Rufus",
        "scraped_at": RUN_DATETIME,
        "total_asins": sum(1 for item in run_statuses.values() if item.get("status") == "ok"),
        "monitored_asins": len(ASINS),
        "failed_asins": failed_asins,
        "statuses": {
            asin: {
                "status": item.get("status", "missing"),
                "error": item.get("error", ""),
                "question_count": len(item.get("questions", [])),
            }
            for asin, item in run_statuses.items()
        },
        "data": deduped_questions,
    }


# ─── 输出 Excel ────────────────────────────────────────────────────────────────

def save_excel(asin_questions: dict, analysis: dict, run_statuses: dict = None):
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)

    def style_header(ws):
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

    total = analysis["total_asins"]
    q2a = analysis["question_to_asins"]

    # Sheet 1：各ASIN完整问题列表（去重后）
    ws1 = wb.active
    ws1.title = "各ASIN问题全览"
    ws1.append(["ASIN", f"{AI_ASSISTANT_SHORT} 问题", "共现ASIN数", "覆盖率", "类型"])
    style_header(ws1)
    for asin, questions in asin_questions.items():
        for q in questions:
            cnt = len(q2a.get(q, []))
            pct = f"{cnt/total:.0%}" if total else "0%"
            if cnt >= analysis["threshold"]:
                qtype = "高频共性"
            elif cnt >= 2:
                qtype = f"中频({cnt}个ASIN)"
            else:
                qtype = "独有"
            ws1.append([asin, q, cnt, pct, qtype])
    ws1.column_dimensions["A"].width = 15
    ws1.column_dimensions["B"].width = 65
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 16

    # Sheet 2：高频共性问题
    ws2 = wb.create_sheet("高频共性问题")
    ws2.append(["问题", f"出现ASIN数（共{total}个）", "覆盖率", "涉及ASIN"])
    style_header(ws2)
    for q, asins in analysis["high_freq"]:
        ws2.append([q, len(asins), f"{len(asins)/total:.0%}", "  |  ".join(asins)])
    ws2.column_dimensions["A"].width = 65
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 50

    # Sheet 3：中频问题
    ws3 = wb.create_sheet("中频问题")
    ws3.append(["问题", "出现ASIN数", "覆盖率", "涉及ASIN"])
    style_header(ws3)
    for q, asins in analysis["mid_freq"]:
        ws3.append([q, len(asins), f"{len(asins)/total:.0%}", "  |  ".join(asins)])
    ws3.column_dimensions["A"].width = 65
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 80

    # Sheet 4：差异性问题（独有）
    ws4 = wb.create_sheet("差异化独有问题")
    ws4.append(["ASIN", "独有问题"])
    style_header(ws4)
    for asin, questions in analysis["unique_questions"].items():
        for q in questions:
            ws4.append([asin, q])
    ws4.column_dimensions["A"].width = 15
    ws4.column_dimensions["B"].width = 65

    # Sheet 5：抓取状态
    if run_statuses:
        ws5 = wb.create_sheet("抓取状态")
        ws5.append(["ASIN", "状态", "问题数", "错误信息"])
        style_header(ws5)
        for asin in ASINS:
            item = run_statuses.get(asin, {})
            ws5.append([
                asin,
                item.get("status", "missing"),
                len(item.get("questions", [])),
                item.get("error", ""),
            ])
        ws5.column_dimensions["A"].width = 15
        ws5.column_dimensions["B"].width = 14
        ws5.column_dimensions["C"].width = 10
        ws5.column_dimensions["D"].width = 60

    wb.save(OUTPUT_EXCEL)
    print(f"\n✅ Excel 已保存: {OUTPUT_EXCEL}")


# ─── 输出 Markdown 报告 ────────────────────────────────────────────────────────

def save_markdown_report(asin_questions: dict, analysis: dict, failed_asins: list,
                         wow_data: dict = None, run_statuses: dict = None) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    total = analysis["total_asins"]
    threshold = analysis["threshold"]
    pct_threshold = threshold / total if total else 0
    success_count = sum(1 for qs in asin_questions.values() if qs)
    failed_str = ", ".join(failed_asins) if failed_asins else "无"
    q2a = analysis["question_to_asins"]

    lines = []
    lines.append(f"# {AI_ASSISTANT_SHORT} 问题分析报告 — 手机三脚架品类\n")
    lines.append(
        f"> 抓取时间：{RUN_DATETIME}　"
        f"监控：{len(ASINS)} 个ASIN　"
        f"成功抓到 {AI_ASSISTANT_SHORT}：{success_count} 个　"
        f"未抓到/失败：{failed_str}\n"
    )
    if run_statuses:
        status_counts = Counter(item.get("status", "missing") for item in run_statuses.values())
        status_summary = "；".join(f"{status}: {count}" for status, count in sorted(status_counts.items()))
        lines.append(f"> 状态统计：{status_summary}。高频/中频覆盖率按成功抓到 {AI_ASSISTANT_SHORT} 的 ASIN 计算。\n")
        if failed_asins:
            lines.append("### 本次未计入统计的 ASIN\n")
            lines.append("| ASIN | 状态 | 说明 |")
            lines.append("|------|------|------|")
            status_labels = {
                "timeout": "页面加载超时或商品主体未完整加载",
                "no_rufus": f"页面已加载，但未识别到 {AI_ASSISTANT_SHORT} 问题区块",
                "blocked": "Amazon 返回封锁/503 页面",
                "captcha": "Amazon 返回验证码页面",
                "not_found": "Amazon 返回页面不存在/不可访问",
                "error": "抓取异常",
            }
            for asin in failed_asins:
                item = run_statuses.get(asin, {})
                status = item.get("status", "missing")
                detail = status_labels.get(status, item.get("error", "") or "未知")
                if item.get("error") and status in {"timeout", "error"}:
                    detail = f"{detail}: {item['error'].splitlines()[0]}"
                lines.append(f"| {asin} | {status} | {detail} |")
            lines.append("")

    lines.append(f"## 零、{AI_ASSISTANT_SHORT} 问题在 Listing 页面中的位置\n")
    lines.append(f"{AI_ASSISTANT_LABEL} 问题出现在 **主图下方、Bullet Points 上方**，同时页面顶部入口已可能显示为「alexa for shopping」。\n")
    lines.append("```")
    lines.append("主图 / 变体选择（颜色/尺寸）")
    lines.append("  ↓")
    lines.append(f"🤖 {AI_ASSISTANT_SHORT}  ← AI 购物助手问题区块（此处）")
    lines.append("  ↓")
    lines.append("About this item（Bullet Points / 五点）")
    lines.append("  ↓")
    lines.append("A+ Content / Product Description")
    lines.append("  ↓")
    lines.append("Customer Reviews")
    lines.append("```\n")
    lines.append("**区块形态：**")
    lines.append("- 标题/入口：可能显示为「Ask Rufus」「Rufus AI」或「alexa for shopping」")
    lines.append("- 展示形式：3~5 个 chip 按钮，横向排列")
    lines.append("- 末尾固定附有「Ask something else」按钮（供用户自由提问）")
    lines.append("- DOM 容器：`id=\"dpx-nice-widget-container\"`，feature name：`nile-inline`\n")
    lines.append("**曝光意义：**")
    lines.append("用户浏览图片后立即看到，位于购买决策流程的最前端，点击率高，是 Amazon 引导用户与 AI 互动的核心入口。Listing 的关键词直接影响此处展示的问题内容（写什么问什么）。\n")
    lines.append("---\n")

    # WoW 对比分析（有上周数据时插入）
    if wow_data:
        lines.append("## 〇、WoW 对比分析\n")
        lines.append(f"> 对比基准：上周抓取时间 {wow_data['prev_scraped_at']}　上周成功 ASIN 数：{wow_data['prev_total_asins']}\n")

        # 高频变化
        lines.append("### 高频共性问题变化\n")
        if wow_data["high_freq_new"]:
            lines.append("**🆕 新进入高频（本周新出现 ≥35% 覆盖）：**")
            for q in wow_data["high_freq_new"]:
                lines.append(f"- {q}")
            lines.append("")
        if wow_data["high_freq_dropped"]:
            lines.append("**📉 退出高频（上周高频、本周覆盖率下降）：**")
            for q in wow_data["high_freq_dropped"]:
                lines.append(f"- {q}")
            lines.append("")
        if wow_data["hf_coverage_change"]:
            lines.append("**📊 高频问题覆盖率变化：**")
            lines.append("| 问题 | 上周 | 本周 | 变化 |")
            lines.append("|------|------|------|------|")
            for item in wow_data["hf_coverage_change"]:
                q = item["question"]
                p_cnt, c_cnt = item["prev"], item["curr"]
                p_total, c_total = item["total_prev"], item["total_curr"]
                delta = c_cnt - p_cnt
                arrow = f"▲{delta}" if delta > 0 else f"▼{abs(delta)}"
                p_pct = f"{p_cnt/p_total:.0%}" if p_total else "0%"
                c_pct = f"{c_cnt/c_total:.0%}" if c_total else "0%"
                lines.append(f"| {q} | {p_cnt}/{p_total} ({p_pct}) | {c_cnt}/{c_total} ({c_pct}) | {arrow} |")
            lines.append("")
        if not wow_data["high_freq_new"] and not wow_data["high_freq_dropped"] and not wow_data["hf_coverage_change"]:
            lines.append("_高频问题无变化_\n")

        # 全量新问题
        if wow_data["brand_new_questions"]:
            lines.append("### 全新问题（本周新增，上周未见）\n")
            for q in wow_data["brand_new_questions"]:
                lines.append(f"- {q}")
            lines.append("")

        # 消失的问题
        if wow_data["disappeared_questions"]:
            lines.append("### 消失的问题（上周有、本周未见）\n")
            for q in wow_data["disappeared_questions"]:
                lines.append(f"- {q}")
            lines.append("")

        # 各ASIN级别变化
        if wow_data["asin_changes"]:
            lines.append("### 各ASIN问题变化详情\n")
            for asin, changes in wow_data["asin_changes"].items():
                lines.append(f"**{asin}**")
                for q in changes["added"]:
                    lines.append(f"- ➕ {q}")
                for q in changes["removed"]:
                    lines.append(f"- ➖ {q}")
                lines.append("")
        else:
            lines.append("### 各ASIN问题变化\n_各 ASIN 问题内容与上周一致_\n")

        lines.append("---\n")

    # 一、高频共性问题
    lines.append("## 一、高频共性问题\n")
    lines.append(f"> 出现在 ≥{threshold} 个ASIN（{pct_threshold:.0%}+）\n")
    lines.append("| 问题 | 出现ASIN数 | 覆盖率 |")
    lines.append("|------|-----------|--------|")
    for q, asins in analysis["high_freq"]:
        lines.append(f"| {q} | {len(asins)}/{total} | {len(asins)/total:.0%} |")
    lines.append("")

    # 二、差异化问题
    lines.append("## 二、差异化问题（各ASIN独有）\n")
    for asin, questions in analysis["unique_questions"].items():
        if questions:
            lines.append(f"**{asin}**")
            for q in questions:
                lines.append(f"- {q}")
            lines.append("")

    # 三、中频问题
    lines.append("## 三、中频问题（2~{max_cnt}个ASIN共有）\n".format(
        max_cnt=threshold - 1
    ))
    lines.append("| 问题 | ASIN数 | 涉及ASIN |")
    lines.append("|------|--------|---------|")
    for q, asins in analysis["mid_freq"]:
        lines.append(f"| {q} | {len(asins)} | {', '.join(asins)} |")
    lines.append("")

    # 四、洞察
    lines.append(f"## 四、{AI_ASSISTANT_SHORT} 逻辑洞察\n")
    lines.append("### 规律总结\n")
    top_questions = analysis["high_freq"][:5]
    if top_questions:
        for i, (q, asins) in enumerate(top_questions, start=1):
            lines.append(f"{i}. **{q}**：覆盖 {len(asins)}/{total} 个成功 ASIN")
    else:
        lines.append(f"1. 本次未形成高频共性问题，说明 {AI_ASSISTANT_SHORT} 问题在不同 listing 间差异较大，或有效样本不足。")
    lines.append("")
    lines.append("### Listing 优化启示\n")
    lines.append("- 高频问题适合检查本品与竞品是否都已在标题、五点、A+ 中明确回应。")
    lines.append("- 独有问题通常对应该 listing 的差异化卖点或特殊表达，值得反查页面文案来源。")
    lines.append(f"- 未抓到 {AI_ASSISTANT_SHORT} 的 ASIN 不参与覆盖率分母，需结合抓取状态判断是页面无 AI 购物助手问题还是访问失败。")
    lines.append("")

    # 五、各ASIN完整问题列表
    lines.append("## 五、各ASIN完整问题列表\n")
    for asin, questions in asin_questions.items():
        lines.append(f"### {asin}\n")
        if not questions:
            lines.append(f"_未抓取到 {AI_ASSISTANT_SHORT} 问题_\n")
            continue
        for q in questions:
            cnt = len(q2a.get(q, []))
            if cnt >= threshold:
                tag = f"共{cnt}个ASIN"
            elif cnt >= 2:
                tag = f"共{cnt}个ASIN"
            else:
                tag = "独有"
            lines.append(f"- [{tag}] {q}")
        lines.append("")

    content = "\n".join(lines)
    OUTPUT_MD.write_text(content, encoding="utf-8")
    print(f"✅ Markdown 报告已保存: {OUTPUT_MD}")
    return OUTPUT_MD


# ─── Google Drive 上传 ─────────────────────────────────────────────────────────

def upload_to_gdrive(file_path: Path, folder_id: str) -> str:
    """上传文件到 Google Drive，返回文件分享 URL"""
    try:
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload

        creds = Credentials.from_authorized_user_file(str(GDRIVE_TOKEN))
        service = build("drive", "v3", credentials=creds)

        mime_map = {
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".md": "text/markdown",
        }
        mime = mime_map.get(file_path.suffix, "application/octet-stream")

        # 检查同名文件是否已存在，存在则更新
        existing = service.files().list(
            q=f"name='{file_path.name}' and '{folder_id}' in parents and trashed=false",
            fields="files(id,name)"
        ).execute().get("files", [])

        if existing:
            file_id = existing[0]["id"]
            media = MediaFileUpload(str(file_path), mimetype=mime)
            service.files().update(fileId=file_id, media_body=media).execute()
            print(f"  ↑ 已更新 Drive 文件: {file_path.name} (id={file_id})")
        else:
            metadata = {"name": file_path.name, "parents": [folder_id]}
            media = MediaFileUpload(str(file_path), mimetype=mime)
            result = service.files().create(
                body=metadata, media_body=media, fields="id"
            ).execute()
            file_id = result["id"]
            print(f"  ↑ 已上传新文件: {file_path.name} (id={file_id})")

        # 设置任何人可查看
        service.permissions().create(
            fileId=file_id,
            body={"type": "anyone", "role": "reader"}
        ).execute()

        return f"https://drive.google.com/file/d/{file_id}/view"

    except Exception as e:
        print(f"⚠ Google Drive 上传失败: {e}")
        return ""


# ─── 飞书通知 ──────────────────────────────────────────────────────────────────

def send_feishu_report(analysis: dict, asin_questions: dict,
                       drive_links: dict = None, wow_data: dict = None,
                       run_statuses: dict = None):
    """
    drive_links = {"md": "https://...", "excel": "https://..."}
    """
    total = analysis["total_asins"]
    high_freq = analysis["high_freq"]
    unique_questions = analysis["unique_questions"]

    asin_lines = []
    for asin, qs in asin_questions.items():
        status = f"✓ {len(qs)} 个问题" if qs else "⚠ 未抓取到"
        asin_lines.append(f"{asin}：{status}")

    if high_freq:
        hf_lines = "\n".join([
            f"{i+1}. {q}（{len(asins)}/{total} 个ASIN）"
            for i, (q, asins) in enumerate(high_freq[:10])
        ])
    else:
        hf_lines = f"无（各 ASIN 问题差异较大，或均未抓到 {AI_ASSISTANT_SHORT} 问题）"

    unique_lines = []
    for asin, qs in unique_questions.items():
        if qs:
            preview = "；".join(qs[:2]) + ("…" if len(qs) > 2 else "")
            unique_lines.append(f"**{asin}** 独有 {len(qs)} 个：{preview}")
    unique_summary = "\n".join(unique_lines) if unique_lines else "无明显独有问题"

    success_count = sum(1 for qs in asin_questions.values() if qs)
    monitored_count = len(run_statuses) if run_statuses else len(ASINS)
    status_counts = Counter(item.get("status", "missing") for item in run_statuses.values()) if run_statuses else Counter()
    failure_summary = ""
    if status_counts:
        parts = [f"{status}: {count}" for status, count in sorted(status_counts.items()) if status != "ok"]
        if parts:
            failure_summary = "\n**未计入原因：** " + "；".join(parts)

    # 构造文件链接文字
    links_content = ""
    if drive_links:
        link_parts = []
        if drive_links.get("md"):
            link_parts.append(f"[📄 分析报告（MD）]({drive_links['md']})")
        if drive_links.get("excel"):
            link_parts.append(f"[📊 {AI_ASSISTANT_SHORT}问题汇总（Excel）]({drive_links['excel']})")
        if link_parts:
            links_content = "**📁 文件链接：**\n" + "　　".join(link_parts)

    elements = [
        {
            "tag": "div",
            "text": {
                "tag": "lark_md",
                "content": (
                    f"**分析时间：** {RUN_DATETIME}\n"
                    f"**监控 ASIN：** {monitored_count} 个　　"
                    f"**成功抓到 {AI_ASSISTANT_SHORT}：** {success_count} 个　　"
                    f"**统计分母：** {total} 个"
                    f"{failure_summary}"
                )
            }
        },
    ]

    if links_content:
        elements.append({"tag": "hr"})
        elements.append({
            "tag": "div",
            "text": {"tag": "lark_md", "content": links_content}
        })

    elements += [
        {"tag": "hr"},
        {
            "tag": "div",
            "text": {
                "tag": "lark_md",
                "content": (
                    f"**🔁 高频共性问题**（≥ {analysis['threshold']} 个ASIN 出现）\n{hf_lines}"
                )
            }
        },
        {"tag": "hr"},
        {
            "tag": "div",
            "text": {
                "tag": "lark_md",
                "content": f"**🔀 差异性问题摘要**\n{unique_summary}"
            }
        },
    ]

    # WoW 对比模块
    if wow_data:
        wow_lines = []
        wow_lines.append(f"📅 对比基准：{wow_data['prev_scraped_at']}")

        if wow_data["high_freq_new"]:
            wow_lines.append("**🆕 新进入高频：**")
            for q in wow_data["high_freq_new"]:
                wow_lines.append(f"- {q}")
        if wow_data["high_freq_dropped"]:
            wow_lines.append("**📉 退出高频：**")
            for q in wow_data["high_freq_dropped"]:
                wow_lines.append(f"- {q}")
        if wow_data["hf_coverage_change"]:
            wow_lines.append("**📊 高频覆盖率变化：**")
            for item in wow_data["hf_coverage_change"]:
                delta = item["curr"] - item["prev"]
                arrow = f"▲{delta}" if delta > 0 else f"▼{abs(delta)}"
                wow_lines.append(
                    f"- {item['question']}：{item['prev']}/{item['total_prev']} → "
                    f"{item['curr']}/{item['total_curr']} ({arrow})"
                )
        if wow_data["brand_new_questions"]:
            wow_lines.append(f"**✨ 全新问题（{len(wow_data['brand_new_questions'])}条）：**")
            for q in wow_data["brand_new_questions"][:5]:
                wow_lines.append(f"- {q}")
            if len(wow_data["brand_new_questions"]) > 5:
                wow_lines.append(f"- …等{len(wow_data['brand_new_questions'])}条，详见报告")
        if wow_data["disappeared_questions"]:
            wow_lines.append(f"**👻 消失的问题（{len(wow_data['disappeared_questions'])}条）：**")
            for q in wow_data["disappeared_questions"][:3]:
                wow_lines.append(f"- {q}")
            if len(wow_data["disappeared_questions"]) > 3:
                wow_lines.append(f"- …等{len(wow_data['disappeared_questions'])}条")

        changed_asins = len(wow_data["asin_changes"])
        if changed_asins:
            wow_lines.append(f"**📝 ASIN 级别变化：** {changed_asins} 个ASIN有问题增减，详见报告")
        else:
            wow_lines.append("**📝 ASIN 级别：** 各ASIN问题与上周一致")

        if not any([wow_data["high_freq_new"], wow_data["high_freq_dropped"],
                    wow_data["hf_coverage_change"], wow_data["brand_new_questions"],
                    wow_data["disappeared_questions"]]):
            wow_lines.append(f"_本周 {AI_ASSISTANT_SHORT} 问题整体无明显变化_")

        elements += [
            {"tag": "hr"},
            {
                "tag": "div",
                "text": {
                    "tag": "lark_md",
                    "content": "**📈 WoW 对比分析**\n" + "\n".join(wow_lines)
                }
            },
        ]

    elements += [
        {"tag": "hr"},
        {
            "tag": "note",
            "elements": [{
                "tag": "plain_text",
                "content": f"报告日期：{RUN_TIMESTAMP} | Excel：{OUTPUT_EXCEL.name}"
            }]
        }
    ]

    payload = {
        "msg_type": "interactive",
        "card": {
            "header": {
                "title": {"tag": "plain_text", "content": f"📊 Amazon {AI_ASSISTANT_SHORT} 问题分析报告"},
                "template": "blue"
            },
            "elements": elements
        }
    }

    try:
        resp = requests.post(
            WEBHOOK_URL,
            headers={"Content-Type": "application/json"},
            data=json.dumps(payload, ensure_ascii=False),
            timeout=10
        )
        resp.raise_for_status()
        result = resp.json()
        if result.get("code") == 0 or result.get("StatusCode") == 0:
            print("✅ 飞书报告已发送")
        else:
            print(f"⚠ 飞书推送返回: {result}")
    except Exception as e:
        print(f"✗ 飞书推送失败: {e}")


# ─── 主流程 ────────────────────────────────────────────────────────────────────

def main():
    print(f"开始抓取 {len(ASINS)} 个 ASIN 的 {AI_ASSISTANT_SHORT} 问题...\n")
    print("⚠ 请确认已开启美区 VPN\n")

    # 加载上周存档（用于 WoW 对比）
    prev_questions = {}
    prev_scraped_at = ""
    archive_file = OUTPUT_DIR / "rufus_archive.json"
    if archive_file.exists():
        try:
            prev_archive = json.loads(archive_file.read_text(encoding="utf-8"))
            prev_run = previous_archive_run(prev_archive, RUN_DATETIME[:10])
            if not prev_run:
                prev_run = latest_archive_run(prev_archive)
            prev_scraped_at = prev_run.get("scraped_at", "")
            prev_raw_questions = prev_run.get("data", {})
            prev_statuses = prev_run.get("statuses", {})
            # 归一化 + 去重，消除存档中的隐藏字符差异
            prev_questions = {
                asin: filter_questions(qs)
                for asin, qs in prev_raw_questions.items()
                if (prev_statuses.get(asin, {}).get("status") == "ok" or (
                    not prev_statuses and qs
                ))
            }
            print(f"✅ 已加载上周存档（{prev_scraped_at}，{len(prev_questions)} 个ASIN）\n")
        except Exception as e:
            print(f"⚠ 加载上周存档失败: {e}\n")

    asin_questions = {}
    run_statuses = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=HEADLESS,
            args=[
                "--no-sandbox",
                "--single-process",
                "--disable-gpu",
                "--disable-blink-features=AutomationControlled",
                "--window-size=1440,900",
            ]
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1440, "height": 900},
            locale="en-US",
            timezone_id="America/New_York",
        )
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', { get: () => undefined });"
        )
        context.route(
            "**/*",
            lambda route, request: (
                route.abort()
                if request.resource_type in BLOCKED_RESOURCE_TYPES
                else route.continue_()
            )
        )

        # 注入 Amazon 登录 cookie
        if COOKIES_FILE.exists():
            cookies = json.loads(COOKIES_FILE.read_text(encoding="utf-8"))
            valid_same_site = {"Strict", "Lax", "None"}
            for c in cookies:
                if c.get("sameSite") not in valid_same_site:
                    c["sameSite"] = "None"
                for key in ["hostOnly", "session", "storeId"]:
                    c.pop(key, None)
            context.add_cookies(cookies)
            print(f"✅ 已注入 Amazon cookies（{len(cookies)} 条）\n")
        else:
            print(f"⚠ 未找到 {COOKIES_FILE}，将以未登录状态访问\n")

        for i, asin in enumerate(ASINS):
            print(f"[{i+1}/{len(ASINS)}] ASIN: {asin}")
            result = scrape_rufus_questions(asin, context)
            run_statuses[asin] = result
            asin_questions[asin] = result["questions"]

            if i < len(ASINS) - 1:
                print(f"  等待 {DELAY_BETWEEN_ASINS}s...\n")
                time.sleep(DELAY_BETWEEN_ASINS)

        retry_transient_failures(context, asin_questions, run_statuses)

        try:
            browser.close()
        except Exception:
            pass

    # 分析
    print("\n─── 分析中 ──────────────────────────────────────────")
    # 当次抓取归一化 + 去重
    deduped_questions = {
        asin: filter_questions(qs)
        for asin, qs in asin_questions.items()
    }
    successful_questions = {
        asin: qs for asin, qs in deduped_questions.items()
        if run_statuses.get(asin, {}).get("status") == "ok" and qs
    }
    analysis = analyze_questions(successful_questions)

    failed_asins = [a for a, item in run_statuses.items() if item.get("status") != "ok"]
    success_count = len(successful_questions)
    not_found_count = sum(1 for item in run_statuses.values() if item.get("status") == "no_rufus")
    archive_quality_ok = success_count >= MIN_SUCCESSFUL_ASINS and not_found_count <= MAX_NOT_FOUND_ASINS
    print(f"\n成功抓取：{success_count}/{len(ASINS)} 个 ASIN")
    if not archive_quality_ok:
        print(
            f"⚠ 数据质量未达归档门槛：成功 {success_count}/{len(ASINS)}，"
            f"未抓到 {not_found_count}；本次不会覆盖 archive。"
        )

    print(f"\n高频共性问题（≥{analysis['threshold']} 个ASIN）：")
    for q, asins in analysis["high_freq"]:
        print(f"  [{len(asins)}/{success_count}] {q}")

    # WoW 对比
    wow_data = None
    if prev_questions:
        print("\n─── WoW 对比分析 ──────────────────────────────────")
        wow_data = compare_wow(successful_questions, prev_questions, analysis, prev_scraped_at)
        if wow_data:
            print(f"  高频新增：{len(wow_data['high_freq_new'])} 条  |  退出：{len(wow_data['high_freq_dropped'])} 条")
            print(f"  全新问题：{len(wow_data['brand_new_questions'])} 条  |  消失：{len(wow_data['disappeared_questions'])} 条")
            print(f"  ASIN 有变化：{len(wow_data['asin_changes'])} 个")

    # 保存本地文件
    save_excel(deduped_questions, analysis, run_statuses=run_statuses)
    save_markdown_report(
        deduped_questions,
        analysis,
        failed_asins,
        wow_data=wow_data,
        run_statuses=run_statuses,
    )

    # 上传 Google Drive
    print("\n─── 上传 Google Drive ────────────────────────────────")
    drive_links = {}
    drive_links["md"] = upload_to_gdrive(OUTPUT_MD, GDRIVE_FOLDER_ID)
    drive_links["excel"] = upload_to_gdrive(OUTPUT_EXCEL, GDRIVE_FOLDER_ID)

    # 飞书通知
    send_feishu_report(
        analysis,
        deduped_questions,
        drive_links=drive_links,
        wow_data=wow_data,
        run_statuses=run_statuses,
    )

    # 存档 JSON（覆盖，下次运行时作为"上周"基准）
    current_run = build_archive_run(deduped_questions, run_statuses)
    previous_history = []
    if archive_file.exists():
        try:
            existing_archive = json.loads(archive_file.read_text(encoding="utf-8"))
            if isinstance(existing_archive.get("history"), list):
                previous_history = existing_archive["history"]
            elif existing_archive.get("data"):
                previous_history = [existing_archive]
        except Exception:
            previous_history = []
    if archive_quality_ok:
        archive = {
            "version": 2,
            "updated_at": RUN_DATETIME,
            "history": (previous_history + [current_run])[-12:],
            # 保留旧字段，方便旧脚本或人工快速查看最近一次结果
            **current_run,
        }
        archive_file.write_text(
            json.dumps(archive, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    else:
        rejected_path = RAW_DIR / f"rejected_archive_run_{RUN_TIMESTAMP}.json"
        rejected_path.write_text(
            json.dumps(current_run, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"⚠ 本次结果已另存为低质量运行记录: {rejected_path}")

    if success_count == 0:
        print(f"\n⚠ 全部 ASIN 未抓到数据，请查看 debug/ 目录。")

    print("\n✅ 全部完成！")


if __name__ == "__main__":
    main()
