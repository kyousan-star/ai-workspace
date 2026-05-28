#!/usr/bin/env python3
"""
Amazon 品类数据准备脚本 — Steps 1 & 3

用法：
  # Step 1：合并多个竞品报告 Excel
  python prepare_data.py --step merge --dir ~/Documents/选品洞察/US/US\ handheld\ massager --category "handheld massager"

  # Step 3：打标完成后，生成季度结构化表格
  python prepare_data.py --step structure --dir ~/Documents/选品洞察/US/US\ handheld\ massager --category "handheld massager" --price-tiers "0-50,50-100,100-150,150-200,200+"

  # 一次运行 Step 1 + Step 3（跳过打标，适合快速验证数据）
  python prepare_data.py --step all --dir ... --category ...

Step 2（维度发现 + 打标）由 Claude 交互完成，脚本不涉及。

依赖：pip install pandas openpyxl xlsxwriter
"""

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ── 列名映射（如源文件列名不同，在此修改） ──────────────────────────────────
COL = {
    # 产品详情 sheet
    "date":         "日期",
    "category":     "品类",
    "parent_asin":  "parentasin",
    "brand":        "品牌",
    "title":        "产品名称",
    "bullets":      "五点描述",
    "main_image":   "主图",
    "monthly_sales": "月销量",       # 或 "月销售量"
    "monthly_revenue": "月销售额",
    "asp":          "ASP",           # 或从销售额/销量计算
    "asin":         "asin",
    # 子体 sheet
    "variant_asin": "变体asin",
    "list_date":    "上架时间",
    "url":          "url",
}

# 月份格式：保证 2024.10 不被存成 2024.1
MONTH_FMT = lambda y, m: f"{y}.{m:02d}"


# ══════════════════════════════════════════════════════════════════════════════
# 工具函数
# ══════════════════════════════════════════════════════════════════════════════

def extract_report_month(sheet_first_row_value: str) -> str | None:
    """
    从 '产品详情' sheet 第一行内容提取报告月份。
    支持格式：
      '2026-02-01至2026-03-01竞争报告'  → '2026.02'
      '2025-11-01至2025-12-01...'       → '2025.11'
    """
    m = re.search(r"(\d{4})-(\d{2})-\d{2}至", str(sheet_first_row_value))
    if m:
        return MONTH_FMT(int(m.group(1)), int(m.group(2)))
    # 备用：从 2026.02 格式直接读
    m2 = re.search(r"(\d{4})\.(\d{2})", str(sheet_first_row_value))
    if m2:
        return MONTH_FMT(int(m2.group(1)), int(m2.group(2)))
    return None


def detect_report_month(xlsx_path: Path) -> str | None:
    """读取 Excel 第一个 sheet 的 A1 单元格，尝试提取报告月份。"""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        # 尝试 "产品详情" 或第一个 sheet
        sheet_names = wb.sheetnames
        target = next((s for s in sheet_names if "产品" in s), sheet_names[0])
        ws = wb[target]
        for row in ws.iter_rows(min_row=1, max_row=3, max_col=5, values_only=True):
            for cell in row:
                if cell:
                    month = extract_report_month(str(cell))
                    if month:
                        wb.close()
                        return month
        wb.close()
    except Exception as e:
        print(f"  ⚠ 无法读取 {xlsx_path.name}: {e}")
    return None


def month_to_quarter(month_str: str) -> str:
    """'2025.03' → '2025Q1'"""
    y, m = month_str.split(".")
    q = (int(m) - 1) // 3 + 1
    return f"{y}Q{q}"


def parse_price_tiers(tiers_str: str) -> list[tuple]:
    """
    '0-50,50-100,100+' → [(0,50), (50,100), (100, inf)]
    支持末尾 '+' 或 '200+' 写法
    """
    result = []
    for t in tiers_str.split(","):
        t = t.strip()
        if "+" in t:
            low = float(t.replace("+", ""))
            result.append((low, float("inf")))
        elif "-" in t:
            parts = t.split("-")
            result.append((float(parts[0]), float(parts[1])))
    return result


def assign_price_tier(price: float, tiers: list[tuple]) -> str:
    """给单个价格打价位段标签。"""
    for low, high in tiers:
        if high == float("inf"):
            if price >= low:
                return f"${int(low)}+"
        else:
            if low <= price < high:
                return f"${int(low)}-{int(high)}"
    return "未知"


# ══════════════════════════════════════════════════════════════════════════════
# Step 1：多表合并
# ══════════════════════════════════════════════════════════════════════════════

def step_merge(data_dir: Path, category: str) -> Path:
    """
    合并目录下所有竞品报告 Excel，输出 us {category} 合并.xlsx
    """
    print(f"\n{'='*60}")
    print(f"Step 1：多表合并")
    print(f"目录：{data_dir}")
    print(f"品类：{category}")
    print(f"{'='*60}")

    xlsx_files = sorted(data_dir.glob("*.xlsx"))
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("us ") and not f.name.startswith("~$")]

    if not xlsx_files:
        print("❌ 目录下没有找到原始 Excel 文件（排除了 'us ' 开头的输出文件）")
        sys.exit(1)

    print(f"\n发现 {len(xlsx_files)} 个文件，开始识别报告月份...")

    # 识别每个文件的月份
    file_month_map: dict[str, Path] = {}
    skipped = []

    for f in xlsx_files:
        month = detect_report_month(f)
        if month:
            if month in file_month_map:
                print(f"  ⚠ 重复月份 {month}：{f.name} 与 {file_month_map[month].name}，保留前者，跳过后者")
                skipped.append(f)
            else:
                file_month_map[month] = f
                print(f"  ✓ {month}  ←  {f.name}")
        else:
            print(f"  ✗ 无法识别月份：{f.name}，跳过")
            skipped.append(f)

    if not file_month_map:
        print("❌ 没有可用文件，退出")
        sys.exit(1)

    months_sorted = sorted(file_month_map.keys())
    print(f"\n有效月份范围：{months_sorted[0]} ~ {months_sorted[-1]}（共 {len(months_sorted)} 个月）")
    if skipped:
        print(f"跳过文件：{[f.name for f in skipped]}")

    # 合并
    detail_frames = []
    variant_frames = []

    for month, fpath in sorted(file_month_map.items()):
        print(f"\n  读取 {month} ({fpath.name})...")
        try:
            xls = pd.ExcelFile(fpath)
        except Exception as e:
            print(f"  ❌ 读取失败：{e}")
            continue

        # ── 产品详情 sheet ────────────────────────────────────────────────
        detail_sheet = next((s for s in xls.sheet_names if "产品" in s and "详情" in s), None)
        if detail_sheet is None:
            detail_sheet = next((s for s in xls.sheet_names if "产品" in s), None)
        if detail_sheet:
            df = xls.parse(detail_sheet, header=1)  # 第1行为标题（第0行是报告说明）
            df.insert(0, COL["date"], month)
            df.insert(1, COL["category"], category)
            detail_frames.append(df)
            print(f"    产品详情：{len(df)} 行，{len(df.columns)} 列")
        else:
            print(f"    ⚠ 未找到 '产品详情' sheet，可用 sheet：{xls.sheet_names}")

        # ── 子体 sheet ────────────────────────────────────────────────────
        variant_sheet = next((s for s in xls.sheet_names if "子体" in s), None)
        if variant_sheet:
            dfv = xls.parse(variant_sheet, header=1)
            dfv.insert(0, COL["date"], month)
            dfv.insert(1, COL["category"], category)
            variant_frames.append(dfv)
            print(f"    子体：{len(dfv)} 行")
        else:
            print(f"    ⚠ 未找到 '子体' sheet")

    if not detail_frames:
        print("❌ 没有读取到任何 '产品详情' 数据")
        sys.exit(1)

    # 合并并输出
    df_detail = pd.concat(detail_frames, ignore_index=True)
    df_variant = pd.concat(variant_frames, ignore_index=True) if variant_frames else pd.DataFrame()

    # 确保日期列格式正确（防止 2024.1 被存成数字）
    df_detail[COL["date"]] = df_detail[COL["date"]].astype(str)
    if not df_variant.empty:
        df_variant[COL["date"]] = df_variant[COL["date"]].astype(str)

    output_path = data_dir / f"us {category} 合并.xlsx"
    with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
        df_detail.to_excel(writer, sheet_name="产品详情", index=False)
        if not df_variant.empty:
            df_variant.to_excel(writer, sheet_name="子体", index=False)

    print(f"\n✅ 合并完成：{output_path}")
    print(f"   产品详情：{len(df_detail)} 行")
    if not df_variant.empty:
        print(f"   子体：{len(df_variant)} 行")
    print(f"\n⏭  下一步：请用 Claude 执行 Step 2（打标），生成 'us {category} 打标后.xlsx'")

    return output_path


# ══════════════════════════════════════════════════════════════════════════════
# Step 3：数据结构整理（父体 & 子体 季度汇总）
# ══════════════════════════════════════════════════════════════════════════════

def _detect_sales_revenue_cols(df: pd.DataFrame) -> tuple[str | None, str | None]:
    """自动探测销量和销售额列名。"""
    sales_candidates = ["月销量", "月销售量", "销量", "Sales"]
    revenue_candidates = ["月销售额", "销售额", "Revenue"]
    sales_col = next((c for c in sales_candidates if c in df.columns), None)
    revenue_col = next((c for c in revenue_candidates if c in df.columns), None)
    return sales_col, revenue_col


def _build_quarterly_pivot(df: pd.DataFrame, group_key: str, tag_cols: list[str],
                            extra_cols: list[str], price_tiers: list[tuple]) -> pd.DataFrame:
    """
    将月度明细转为季度宽表。
    group_key: 'parentasin' 或 'asin'
    tag_cols:  打标维度列（插在 group_key 后）
    extra_cols: 非聚合的附加列（取最新值）
    """
    date_col = COL["date"]
    sales_col, revenue_col = _detect_sales_revenue_cols(df)

    if sales_col is None:
        print("  ❌ 未找到销量列，请检查 COL['monthly_sales'] 映射")
        return pd.DataFrame()

    # 生成季度列
    df = df.copy()
    df["_quarter"] = df[date_col].apply(month_to_quarter)

    # 所有月份
    all_months = sorted(df[date_col].unique())
    all_quarters = sorted(df["_quarter"].unique())

    # 识别年份
    years = sorted(set(m.split(".")[0] for m in all_months))
    if len(years) < 2:
        prev_year, curr_year = years[0], years[0]
    else:
        prev_year, curr_year = years[0], years[-1]

    # 同期月份：两年中月份数相同
    prev_months = [m for m in all_months if m.startswith(prev_year)]
    curr_months_all = [m for m in all_months if m.startswith(curr_year)]
    # 同期取与 prev_year 相同数量的月份（从最早开始）
    n = len(prev_months)
    curr_same_months = curr_months_all[:n]

    prev_quarters = sorted(set(month_to_quarter(m) for m in prev_months))
    curr_quarters = sorted(set(month_to_quarter(m) for m in curr_months_all))

    def safe_q(q):
        return q.replace(" ", "")

    # 价位段
    asp_col = None
    if revenue_col and sales_col:
        df["_asp"] = df[revenue_col] / df[sales_col].replace(0, pd.NA)
    if "ASP" in df.columns:
        asp_col = "ASP"
    elif "_asp" in df.columns:
        asp_col = "_asp"

    if asp_col and price_tiers:
        df["价位段"] = df[asp_col].apply(
            lambda x: assign_price_tier(float(x), price_tiers) if pd.notna(x) else "未知"
        )
        if "价位段" not in tag_cols:
            tag_cols = ["价位段"] + [c for c in tag_cols if c != "价位段"]

    # 汇总每个 group_key + 季度 的销量/销售额
    agg_dict = {sales_col: "sum"}
    if revenue_col:
        agg_dict[revenue_col] = "sum"

    pivot = df.groupby([group_key, "_quarter"]).agg(agg_dict).reset_index()
    pivot_wide = pivot.pivot(index=group_key, columns="_quarter", values=list(agg_dict.keys()))
    pivot_wide.columns = [f"{col}_{q}" for col, q in pivot_wide.columns]
    pivot_wide = pivot_wide.reset_index()

    # 年度汇总
    for yr, qlist in [(prev_year, prev_quarters), (curr_year, curr_quarters)]:
        sales_qs = [f"{sales_col}_{q}" for q in qlist if f"{sales_col}_{q}" in pivot_wide.columns]
        pivot_wide[f"{yr}Y销量"] = pivot_wide[sales_qs].sum(axis=1) if sales_qs else 0
        if revenue_col:
            rev_qs = [f"{revenue_col}_{q}" for q in qlist if f"{revenue_col}_{q}" in pivot_wide.columns]
            pivot_wide[f"{yr}Y销售额"] = pivot_wide[rev_qs].sum(axis=1) if rev_qs else 0

    # 同期销量
    same_qs = sorted(set(month_to_quarter(m) for m in curr_same_months))
    same_sales_qs = [f"{sales_col}_{q}" for q in same_qs if f"{sales_col}_{q}" in pivot_wide.columns]
    pivot_wide[f"{curr_year}Y同期销量"] = pivot_wide[same_sales_qs].sum(axis=1) if same_sales_qs else 0
    if revenue_col:
        same_rev_qs = [f"{revenue_col}_{q}" for q in same_qs if f"{revenue_col}_{q}" in pivot_wide.columns]
        pivot_wide[f"{curr_year}Y同期销售额"] = pivot_wide[same_rev_qs].sum(axis=1) if same_rev_qs else 0

    # 附加列（取最新值）
    if extra_cols:
        latest = df.sort_values(date_col).groupby(group_key)[extra_cols].last().reset_index()
        pivot_wide = pivot_wide.merge(latest, on=group_key, how="left")

    # 打标维度（取最新值）
    if tag_cols:
        tag_available = [c for c in tag_cols if c in df.columns]
        if tag_available:
            tags = df.sort_values(date_col).groupby(group_key)[tag_available].last().reset_index()
            pivot_wide = pivot_wide.merge(tags, on=group_key, how="left")

    # 品牌列
    if COL["brand"] in df.columns:
        brand_map = df.sort_values(date_col).groupby(group_key)[COL["brand"]].last()
        pivot_wide = pivot_wide.merge(brand_map.rename(COL["brand"]), on=group_key, how="left")

    # 品类列
    pivot_wide.insert(0, COL["category"], category)

    # 列排序：品类 → group_key → 品牌 → 打标维度 → 季度销量 → 年度汇总 → 季度销售额 → ...
    front_cols = [COL["category"], group_key]
    if COL["brand"] in pivot_wide.columns:
        front_cols.append(COL["brand"])
    tag_available = [c for c in tag_cols if c in pivot_wide.columns]
    front_cols += tag_available

    sales_q_cols = [f"{sales_col}_{q}" for q in sorted(all_quarters) if f"{sales_col}_{q}" in pivot_wide.columns]
    y_sales_cols = [f"{prev_year}Y销量", f"{curr_year}Y销量", f"{curr_year}Y同期销量"]
    y_sales_cols = [c for c in y_sales_cols if c in pivot_wide.columns]

    rev_q_cols, y_rev_cols = [], []
    if revenue_col:
        rev_q_cols = [f"{revenue_col}_{q}" for q in sorted(all_quarters) if f"{revenue_col}_{q}" in pivot_wide.columns]
        y_rev_cols = [f"{prev_year}Y销售额", f"{curr_year}Y销售额", f"{curr_year}Y同期销售额"]
        y_rev_cols = [c for c in y_rev_cols if c in pivot_wide.columns]

    extra_available = [c for c in extra_cols if c in pivot_wide.columns]

    ordered_cols = front_cols + sales_q_cols + y_sales_cols + rev_q_cols + y_rev_cols + extra_available
    remaining = [c for c in pivot_wide.columns if c not in ordered_cols]
    pivot_wide = pivot_wide[ordered_cols + remaining]

    return pivot_wide


def step_structure(data_dir: Path, category: str, price_tiers_str: str) -> Path:
    """
    Step 3：读取打标后的文件，生成 us {category} 数据结构整理.xlsx
    """
    print(f"\n{'='*60}")
    print(f"Step 3：数据结构整理")
    print(f"{'='*60}")

    tagged_path = data_dir / f"us {category} 打标后.xlsx"
    if not tagged_path.exists():
        # 如果没有打标文件，尝试用合并文件
        tagged_path = data_dir / f"us {category} 合并.xlsx"
        if tagged_path.exists():
            print(f"⚠ 未找到打标后文件，使用合并文件（无打标维度）：{tagged_path.name}")
        else:
            print(f"❌ 未找到输入文件：us {category} 打标后.xlsx 或 us {category} 合并.xlsx")
            sys.exit(1)

    print(f"读取：{tagged_path.name}")

    try:
        df_detail = pd.read_excel(tagged_path, sheet_name="产品详情")
        print(f"  产品详情：{len(df_detail)} 行 × {len(df_detail.columns)} 列")
    except Exception as e:
        print(f"❌ 读取产品详情失败：{e}")
        sys.exit(1)

    try:
        df_variant = pd.read_excel(tagged_path, sheet_name="子体")
        print(f"  子体：{len(df_variant)} 行")
    except Exception:
        df_variant = pd.DataFrame()
        print("  ⚠ 未找到子体 sheet")

    price_tiers = parse_price_tiers(price_tiers_str) if price_tiers_str else []

    # 识别打标维度列（不含系统列）
    system_cols = {COL["date"], COL["category"], COL["parent_asin"], COL["brand"],
                   COL["title"], COL["bullets"], COL["main_image"],
                   COL["monthly_sales"], COL["monthly_revenue"], COL["asin"],
                   "月销量", "月销售量", "月销售额", "销量", "销售额", "ASP", "_quarter", "_asp"}
    tag_cols = [c for c in df_detail.columns
                if c not in system_cols and not str(c).startswith("_")
                and c not in (COL["date"], COL["category"])]
    # 排除数值型大列（季度数据等）
    tag_cols = [c for c in tag_cols
                if df_detail[c].dtype == object or str(df_detail[c].dtype) == "category"]

    print(f"\n识别到打标维度列：{tag_cols if tag_cols else '（无，将跳过维度）'}")

    # 父体整理
    parent_col = COL["parent_asin"]
    if parent_col not in df_detail.columns:
        # 尝试大小写变体
        parent_col = next((c for c in df_detail.columns
                           if c.lower().replace(" ", "") == "parentasin"), None)
    if parent_col is None:
        print("❌ 未找到 parentasin 列，请检查列名映射")
        sys.exit(1)

    print(f"\n正在生成父体整理...")
    df_parent = _build_quarterly_pivot(
        df_detail, parent_col, tag_cols,
        extra_cols=[],
        price_tiers=price_tiers
    )
    print(f"  父体整理：{len(df_parent)} 行")

    # 子体整理
    df_child = pd.DataFrame()
    if not df_variant.empty:
        asin_col = COL["asin"]
        variant_col = COL["variant_asin"]
        # 检查变体 asin 列
        if variant_col not in df_variant.columns:
            variant_col = next((c for c in df_variant.columns if "asin" in c.lower()), None)
        if variant_col:
            # 同步打标维度：通过 parentasin 映射
            if tag_cols and parent_col in df_variant.columns:
                tag_map = df_detail[[parent_col] + tag_cols].drop_duplicates(subset=[parent_col])
                df_variant = df_variant.merge(tag_map, on=parent_col, how="left", suffixes=("", "_tag"))
                for tc in tag_cols:
                    if f"{tc}_tag" in df_variant.columns:
                        df_variant[tc] = df_variant[tc].fillna(df_variant[f"{tc}_tag"])
                        df_variant.drop(columns=[f"{tc}_tag"], inplace=True)

            print(f"\n正在生成子体整理...")
            extra = [c for c in [COL["url"], COL["list_date"]] if c in df_variant.columns]
            df_child = _build_quarterly_pivot(
                df_variant, variant_col, tag_cols,
                extra_cols=extra,
                price_tiers=price_tiers
            )
            # 补充上架天数
            if COL["list_date"] in df_child.columns:
                import datetime
                df_child["上架天数"] = (
                    pd.Timestamp.today() - pd.to_datetime(df_child[COL["list_date"]], errors="coerce")
                ).dt.days
            print(f"  子体整理：{len(df_child)} 行")

    # 输出
    output_path = data_dir / f"us {category} 数据结构整理.xlsx"
    with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
        if not df_parent.empty:
            df_parent.to_excel(writer, sheet_name="父体整理", index=False)
        if not df_child.empty:
            df_child.to_excel(writer, sheet_name="子体整理", index=False)

    print(f"\n✅ 结构整理完成：{output_path}")
    if not df_parent.empty:
        print(f"   父体整理：{len(df_parent)} 行 × {len(df_parent.columns)} 列")
    if not df_child.empty:
        print(f"   子体整理：{len(df_child)} 行 × {len(df_child.columns)} 列")
    print(f"\n⏭  下一步：将此文件交给 Claude 执行分析模块 A-H")

    return output_path


# ══════════════════════════════════════════════════════════════════════════════
# 入口
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Amazon 品类数据准备脚本（Step 1 合并 + Step 3 结构整理）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument("--step", choices=["merge", "structure", "all"],
                        default="merge", help="执行步骤：merge / structure / all")
    parser.add_argument("--dir", required=True, help="数据目录路径")
    parser.add_argument("--category", required=True, help="品类英文名（如 handheld massager）")
    parser.add_argument("--price-tiers",
                        default="0-50,50-100,100-150,150-200,200+",
                        help="价位段划分，逗号分隔（默认：0-50,50-100,100-150,150-200,200+）")

    args = parser.parse_args()

    data_dir = Path(args.dir).expanduser().resolve()
    if not data_dir.exists():
        print(f"❌ 目录不存在：{data_dir}")
        sys.exit(1)

    if args.step in ("merge", "all"):
        step_merge(data_dir, args.category)

    if args.step in ("structure", "all"):
        step_structure(data_dir, args.category, args.price_tiers)


if __name__ == "__main__":
    main()
